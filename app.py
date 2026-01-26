"""
프로젝트명: 스마트 의류 요척 산출서 (Smart Fabric Yield Calculator)
버전: V36 (Final Clean)
설명: DXF 패턴 파일을 분석하여 원단 소요량(요척)을 자동 산출하는 웹 애플리케이션
주요기능:
  - DXF 파일 파싱 및 형상 분석 (ezdxf, shapely)
  - 패턴 썸네일 그리드 및 인터랙티브 뷰어 (matplotlib, plotly)
  - 원단명/수량 일괄 수정 및 개별 요척 계산
  - Streamlit 기반의 반응형 UI
"""

import streamlit as st
import ezdxf
from shapely.geometry import LineString, Polygon, Point
from shapely import affinity
from shapely.ops import polygonize
import math
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import pandas as pd
import tempfile
import os
import io
import base64

# 상수 임포트
from constants import (
    FABRIC_MAP, FABRIC_COLORS, DEFAULT_FABRIC_COLOR, DEFAULT_FABRIC_NAME,
    SIZE_ORDER, SIZE_ORDER_LIST, MIN_PATTERN_AREA, BASE_SIZE_PREFIXES,
    GRAINLINE_KEYWORDS, GRAINLINE_LAYER_NUMBERS, UNIT_SCALE_INCH_TO_MM,
    get_fabric_color, size_sort_key, get_fabric_name
)

# 네스팅 엔진 임포트
from nesting_engine import NestingEngine, create_nesting_visualization

# 형상 시그니처 DB (수량 추천/학습)
try:
    from pattern_db import PatternDB
    PATTERN_DB_AVAILABLE = True
except ImportError:
    PATTERN_DB_AVAILABLE = False

# Sparrow 네스팅 (State-of-the-art)
try:
    import spyrrow
    SPARROW_AVAILABLE = True
except ImportError:
    SPARROW_AVAILABLE = False


def extract_english_size(size_val):
    """
    사이즈 문자열에서 영문/숫자 부분만 추출
    예: "S축적용" → "S", "XL축적용" → "XL", "M축적용" → "M"
    """
    import re
    if not size_val:
        return size_val
    # 앞부분의 영문/숫자만 추출 (0X, XL, S, M, L, 2XL, 85, 90 등)
    match = re.match(r'^([A-Za-z0-9]+)', size_val.strip())
    if match:
        return match.group(1)
    return size_val


def run_sparrow_nesting(pattern_data, width_cm, time_limit, allow_rotation, spacing, allow_mirror=False, buffer_mm=0, allow_90_rotation=False, seed=42, pattern_order="default"):
    """
    Sparrow 네스팅 실행

    Args:
        pattern_data: 패턴 데이터 리스트 [{coords_cm, quantity, pattern_id, area_cm2}, ...]
        width_cm: 원단 폭 (cm)
        time_limit: 최적화 시간 제한 (초)
        allow_rotation: 180도 회전 허용 여부
        spacing: 패턴 간격 (mm) - cm로 변환하여 적용
        allow_mirror: 뒤집기(좌우 미러링) 허용 여부
        buffer_mm: 패턴 둘레 버퍼 (mm) - 패턴 외곽으로 확장하여 블로킹
        allow_90_rotation: 90도 회전 허용 여부
        seed: 랜덤 시드 (다른 배치 결과 생성)
        pattern_order: 패턴 입력 순서 ("default", "large_first", "small_first", "random")

    Returns:
        네스팅 결과 딕셔너리
    """
    # 패턴 순서 정렬
    import random
    sorted_pattern_data = pattern_data.copy()
    if pattern_order == "large_first":
        sorted_pattern_data.sort(key=lambda p: p.get('area_cm2', 0), reverse=True)
    elif pattern_order == "small_first":
        sorted_pattern_data.sort(key=lambda p: p.get('area_cm2', 0))
    elif pattern_order == "random":
        random.seed(seed)
        random.shuffle(sorted_pattern_data)

    # spyrrow Item 생성
    items = []
    total_area = 0
    item_idx = 0
    buffer_cm = buffer_mm / 10  # mm -> cm

    # 원본 좌표 저장 (버퍼 시각화용)
    original_shapes = {}  # id -> 원본 좌표 (Sparrow 좌표계)
    grainline_data = {}  # id -> 그레인라인 좌표 (Sparrow 좌표계)

    for p in sorted_pattern_data:
        coords = list(p['coords_cm'])
        # 닫힌 폴리곤으로 변환
        if coords[0] != coords[-1]:
            coords = coords + [coords[0]]

        # 원본 좌표 저장 (버퍼 적용 전)
        original_coords = coords.copy()

        # 패턴별 상하좌우 버퍼 (mm -> cm)
        buf_top_cm = p.get('buffer_top', 0) / 10
        buf_bottom_cm = p.get('buffer_bottom', 0) / 10
        buf_left_cm = p.get('buffer_left', 0) / 10
        buf_right_cm = p.get('buffer_right', 0) / 10
        has_directional_buffer = (buf_top_cm > 0 or buf_bottom_cm > 0 or buf_left_cm > 0 or buf_right_cm > 0)

        # 버퍼 적용 (패턴별 상하좌우 또는 전역 둘레 버퍼)
        from shapely.geometry import Polygon as ShapelyPolygon, box
        from shapely.ops import unary_union
        poly = ShapelyPolygon(coords)

        if has_directional_buffer:
            # 상하좌우 개별 버퍼 적용 (affine 변환 사용)
            # 패턴을 상하좌우 방향으로 확장
            from shapely import affinity

            minx, miny, maxx, maxy = poly.bounds
            width = maxx - minx
            height = maxy - miny

            if width > 0 and height > 0:
                # 중심점 계산
                cx = (minx + maxx) / 2
                cy = (miny + maxy) / 2

                # 확장 비율 계산 (버퍼를 포함한 새로운 크기)
                new_width = width + buf_left_cm + buf_right_cm
                new_height = height + buf_top_cm + buf_bottom_cm
                scale_x = new_width / width
                scale_y = new_height / height

                # 스케일 적용 (중심 기준)
                scaled_poly = affinity.scale(poly, xfact=scale_x, yfact=scale_y, origin=(cx, cy))

                # 비대칭 버퍼 보정 (좌우/상하 버퍼가 다를 경우 이동)
                offset_x = (buf_right_cm - buf_left_cm) / 2
                offset_y = (buf_top_cm - buf_bottom_cm) / 2
                buffered_poly = affinity.translate(scaled_poly, xoff=offset_x, yoff=offset_y)

                if buffered_poly.is_valid and not buffered_poly.is_empty:
                    if buffered_poly.geom_type == 'Polygon':
                        coords = list(buffered_poly.exterior.coords)
                    else:
                        # MultiPolygon인 경우 가장 큰 폴리곤 사용
                        largest = max(buffered_poly.geoms, key=lambda g: g.area)
                        coords = list(largest.exterior.coords)
        elif buffer_cm > 0:
            # 전역 둘레 버퍼 (기존 방식)
            buffered_poly = poly.buffer(buffer_cm, join_style=2)  # join_style=2: mitre (각진 모서리)
            if buffered_poly.is_valid and not buffered_poly.is_empty:
                coords = list(buffered_poly.exterior.coords)

        # 좌표 변환: DXF(X=폭, Y=길이) → Sparrow(X=길이, Y=폭)
        # 입력시 X와 Y를 교환하여 Sparrow가 올바르게 처리하도록 함
        swapped_coords = [(y, x) for x, y in coords]
        swapped_original = [(y, x) for x, y in original_coords]  # 원본도 변환

        # 좌우 미러링된 좌표 생성 (Y축 반전) - 좌/우 패턴 짝 배치용
        # Sparrow 좌표계: X=마카길이, Y=원단폭
        # Y축 반전 = 원단폭 방향으로 뒤집기 = 좌우 마주보기
        ys = [c[1] for c in swapped_coords]
        center_y = (min(ys) + max(ys)) / 2
        mirrored_coords = [(x, 2 * center_y - y) for x, y in swapped_coords]
        mirrored_original = [(x, 2 * center_y - y) for x, y in swapped_original]  # 원본도 미러링

        # 그레인라인 좌표 변환 (Sparrow 좌표계)
        grainline_cm = p.get('grainline_cm')
        swapped_grainline = None
        mirrored_grainline = None
        if grainline_cm:
            gl_start, gl_end = grainline_cm
            # DXF(X,Y) → Sparrow(Y,X) 좌표 교환
            swapped_grainline = ((gl_start[1], gl_start[0]), (gl_end[1], gl_end[0]))
            # 미러링된 그레인라인 (Y축 반전)
            mirrored_grainline = (
                (gl_start[1], 2 * center_y - gl_start[0]),
                (gl_end[1], 2 * center_y - gl_end[0])
            )

        # 수량만큼 아이템 생성
        for q in range(p['quantity']):
            unique_id = f"{p['pattern_id']}_{item_idx}"

            # 회전 옵션 설정
            orientations = [0]
            if allow_rotation:
                orientations.append(180)
            if allow_90_rotation:
                orientations.extend([90, 270])

            # 좌우 마주 보기 옵션 (180도 회전과 독립)
            if allow_mirror:
                use_coords = mirrored_coords if (q % 2 == 1) else swapped_coords
                use_original = mirrored_original if (q % 2 == 1) else swapped_original
                use_grainline = mirrored_grainline if (q % 2 == 1) else swapped_grainline
            else:
                use_coords = swapped_coords
                use_original = swapped_original
                use_grainline = swapped_grainline

            # 원본 좌표 저장
            original_shapes[unique_id] = use_original
            # 그레인라인 좌표 저장
            if use_grainline:
                grainline_data[unique_id] = use_grainline

            item = spyrrow.Item(
                id=unique_id,
                shape=use_coords,
                demand=1,
                allowed_orientations=orientations
            )
            items.append(item)
            total_area += p['area_cm2']
            item_idx += 1

    # StripPackingInstance 생성
    spacing_cm = spacing / 10  # mm -> cm
    instance = spyrrow.StripPackingInstance(
        name="nesting",
        strip_height=width_cm,
        items=items
    )

    # 솔버 설정
    config = spyrrow.StripPackingConfig(
        total_computation_time=time_limit,
        min_items_separation=spacing_cm if spacing_cm > 0 else None,
        seed=seed
    )

    # 실행
    solution = instance.solve(config)

    # 결과 변환
    used_length_cm = solution.width
    efficiency = solution.density * 100

    # 배치 정보 추출
    # PlacedItem: id, rotation, translation (x, y)
    placements = []
    item_shapes = {item.id: item.shape for item in items}  # ID -> 버퍼 좌표 매핑

    for placed in solution.placed_items:
        buffered_shape = item_shapes.get(placed.id, [])
        original_shape = original_shapes.get(placed.id, [])  # 원본 좌표 (버퍼 전)
        rotation = placed.rotation
        tx, ty = placed.translation

        # 180도 회전 비허용 시: 180도 회전된 패턴을 원래대로 되돌림
        # spyrrow가 allowed_orientations=[0]을 무시하는 경우 대응
        if not allow_rotation and abs(rotation - 180) < 1:
            rotation = 0
            # 180도 회전 취소를 위해 translation 보정
            # 패턴 중심 기준으로 180도 역회전 필요
            xs = [c[0] for c in buffered_shape]
            ys = [c[1] for c in buffered_shape]
            cx = (min(xs) + max(xs)) / 2
            cy = (min(ys) + max(ys)) / 2
            # 원래 회전 중심에서의 오프셋 계산 후 역보정
            tx = tx + 2 * cx
            ty = ty + 2 * cy

        # 회전 및 이동 적용하여 최종 좌표 계산
        # Sparrow 좌표계 (입력 교환 후): X=마카길이, Y=원단폭
        # 출력시 다시 교환: X=원단폭, Y=마카길이
        cos_r = math.cos(math.radians(rotation))
        sin_r = math.sin(math.radians(rotation))

        # 버퍼 적용된 좌표 변환
        transformed_coords = []
        for x, y in buffered_shape:
            # 회전 (원점 기준)
            rx = x * cos_r - y * sin_r
            ry = x * sin_r + y * cos_r
            # 이동
            px = rx + tx
            py = ry + ty
            # 좌표 교환: Sparrow(X=길이, Y=폭) → 시각화(X=폭, Y=길이)
            transformed_coords.append((py, px))

        # 원본 좌표 변환 (버퍼 전)
        original_transformed = []
        for x, y in original_shape:
            # 회전 (원점 기준)
            rx = x * cos_r - y * sin_r
            ry = x * sin_r + y * cos_r
            # 이동
            px = rx + tx
            py = ry + ty
            # 좌표 교환: Sparrow(X=길이, Y=폭) → 시각화(X=폭, Y=길이)
            original_transformed.append((py, px))

        # 그레인라인 좌표 변환
        grainline_transformed = None
        if placed.id in grainline_data:
            gl_start, gl_end = grainline_data[placed.id]
            transformed_gl = []
            for x, y in [gl_start, gl_end]:
                # 회전 (원점 기준)
                rx = x * cos_r - y * sin_r
                ry = x * sin_r + y * cos_r
                # 이동
                px = rx + tx
                py = ry + ty
                # 좌표 교환: Sparrow(X=길이, Y=폭) → 시각화(X=폭, Y=길이)
                transformed_gl.append((py, px))
            grainline_transformed = (transformed_gl[0], transformed_gl[1])

        placements.append({
            'pattern_id': placed.id,
            'x': ty,  # Sparrow Y → 시각화 X (폭 방향)
            'y': tx,  # Sparrow X → 시각화 Y (길이 방향)
            'rotation': rotation,
            'coords': transformed_coords,  # 버퍼 적용된 좌표
            'original_coords': original_transformed,  # 원본 좌표 (버퍼 전)
            'grainline': grainline_transformed  # 그레인라인 좌표
        })

    return {
        'success': True,
        'placed_count': len(placements),
        'total_count': len(items),
        'used_length': used_length_cm * 10,  # cm -> mm (호환성)
        'used_length_mm': used_length_cm * 10,  # mm (시각화용)
        'used_length_cm': used_length_cm,
        'used_length_yd': used_length_cm / 100 * 1.09361,
        'efficiency': round(efficiency, 1),
        'placements': placements,
        'sparrow_mode': True,
        'buffer_mm': buffer_mm  # 버퍼 크기 저장
    }


def create_sparrow_visualization(result, sheet_width_cm):
    """Sparrow 네스팅 결과 시각화"""
    import matplotlib.pyplot as plt
    import matplotlib.patches as patches
    from matplotlib.path import Path
    import matplotlib.font_manager as fm
    import platform

    # 한글 폰트 설정 (크로스 플랫폼)
    if platform.system() == 'Windows':
        plt.rcParams['font.family'] = 'Malgun Gothic'
    else:
        # Linux: NanumGothic 우선, 없으면 DejaVu Sans
        available_fonts = [f.name for f in fm.fontManager.ttflist]
        if 'NanumGothic' in available_fonts:
            plt.rcParams['font.family'] = 'NanumGothic'
        elif 'Nanum Gothic' in available_fonts:
            plt.rcParams['font.family'] = 'Nanum Gothic'
        else:
            plt.rcParams['font.family'] = 'DejaVu Sans'
    plt.rcParams['axes.unicode_minus'] = False

    if not result.get('success'):
        return None

    fig_width = 12
    used_length_cm = result['used_length_cm']
    aspect = used_length_cm / sheet_width_cm
    fig_height = max(3, min(15, fig_width * aspect * 0.4))

    fig, ax = plt.subplots(figsize=(fig_width, fig_height))

    # 시트 배경
    sheet = patches.Rectangle(
        (0, 0), sheet_width_cm, used_length_cm,
        linewidth=2, edgecolor='green', facecolor='#e8f5e9', alpha=0.5
    )
    ax.add_patch(sheet)

    # 색상 팔레트
    colors = plt.cm.Set3(range(12))

    # 사이즈 개수 확인
    all_sizes = st.session_state.get('all_sizes', [])
    selected_sizes = st.session_state.get('selected_sizes', all_sizes)
    has_multiple_sizes = len(selected_sizes) >= 2

    if has_multiple_sizes:
        # 사이즈 2개 이상: 같은 사이즈는 같은 색상
        # pattern_id에 저장된 사이즈는 4자로 잘려있으므로 앞 4자로 매핑
        size_color_map = {}
        for i, size in enumerate(selected_sizes):
            size_color_map[size] = colors[i % len(colors)]
            size_color_map[size[:4]] = colors[i % len(colors)]  # 잘린 버전도 매핑
    else:
        # 사이즈 1개: 같은 패턴은 같은 색상
        unique_patterns = list(set(p['pattern_id'].split('\n')[0] for p in result['placements']))
        pattern_color_map = {name: colors[i % len(colors)] for i, name in enumerate(unique_patterns)}

    # 패턴 그리기
    for i, p in enumerate(result['placements']):
        coords = p['coords']  # 버퍼 적용된 좌표
        original_coords = p.get('original_coords', [])  # 원본 좌표

        if coords:
            # coords는 이미 배치된 좌표
            xs = [c[0] for c in coords]
            ys = [c[1] for c in coords]

            # 색상 결정
            if has_multiple_sizes:
                # 사이즈별 색상 (pattern_id 형식: "패턴명\n사이즈_인덱스")
                parts = p['pattern_id'].split('\n')
                if len(parts) > 1:
                    # "사이즈_인덱스"에서 사이즈만 추출 (마지막 _인덱스 제거)
                    size_with_idx = parts[1].strip()
                    size_part = size_with_idx.rsplit('_', 1)[0] if '_' in size_with_idx else size_with_idx
                else:
                    size_part = ''
                # 매핑에서 색상 찾기
                color = size_color_map.get(size_part, colors[i % len(colors)])
            else:
                # 패턴별 색상
                pattern_name = p['pattern_id'].split('\n')[0]
                color = pattern_color_map.get(pattern_name, colors[0])

            # 버퍼 적용 여부 확인 (원본 좌표가 있고 버퍼 좌표와 다르면 버퍼 적용됨)
            has_buffer = original_coords and coords != original_coords

            # 버퍼가 있으면 버퍼 영역(실선 테두리) + 원본 패턴(진한 색) 표시
            if has_buffer:
                # 버퍼 영역 (실선 테두리, 연한 배경)
                ax.fill(xs, ys, alpha=0.2, facecolor=color, edgecolor='#333333', linewidth=1.0, linestyle='-')

                # 원본 패턴 (진한 색, 실선 테두리)
                orig_xs = [c[0] for c in original_coords]
                orig_ys = [c[1] for c in original_coords]
                ax.fill(orig_xs, orig_ys, alpha=0.7, facecolor=color, edgecolor='black', linewidth=0.5)
            else:
                # 버퍼 없으면 기존대로 표시
                ax.fill(xs, ys, alpha=0.7, facecolor=color, edgecolor='black', linewidth=0.5)

            # 패턴 ID 표시 (중심점 계산 - Shapely centroid 사용)
            from shapely.geometry import Polygon as ShapelyPoly
            try:
                poly_shape = ShapelyPoly(coords)
                centroid = poly_shape.centroid
                cx, cy = centroid.x, centroid.y
            except:
                cx = sum(xs) / len(xs)
                cy = sum(ys) / len(ys)
            # 패턴ID에서 패턴이름 + 사이즈 표시
            # pattern_id 형식: "인덱스:패턴이름\n사이즈_수량인덱스"
            raw_id = p['pattern_id']
            parts = raw_id.split('\n')
            # 첫 번째 줄에서 "인덱스:" 제거하고 패턴이름 추출
            first_line = parts[0]
            if ':' in first_line:
                pattern_name = first_line.split(':', 1)[1]
            else:
                pattern_name = first_line

            # 사이즈 추출 (두 번째 줄에서 _수량인덱스 제거)
            if len(parts) > 1:
                size_with_idx = parts[1].strip()
                size_part = size_with_idx.rsplit('_', 1)[0] if '_' in size_with_idx else size_with_idx
                label = f"{pattern_name}\n{size_part}" if pattern_name else size_part
            else:
                label = pattern_name
            ax.text(cx, cy, label, ha='center', va='center', fontsize=4, fontweight='bold')

            # 그레인라인 표시 (검정 실선, 크기 50%)
            grainline = p.get('grainline')
            if grainline:
                gl_start, gl_end = grainline
                # 그레인라인 크기를 50%로 축소 (중심점 기준)
                cx_gl = (gl_start[0] + gl_end[0]) / 2
                cy_gl = (gl_start[1] + gl_end[1]) / 2
                gl_start_scaled = (cx_gl + (gl_start[0] - cx_gl) * 0.5, cy_gl + (gl_start[1] - cy_gl) * 0.5)
                gl_end_scaled = (cx_gl + (gl_end[0] - cx_gl) * 0.5, cy_gl + (gl_end[1] - cy_gl) * 0.5)
                ax.plot([gl_start_scaled[0], gl_end_scaled[0]], [gl_start_scaled[1], gl_end_scaled[1]],
                        'k-', linewidth=0.4, alpha=0.8)
                # 화살표 머리 (끝점에)
                dx = gl_end_scaled[0] - gl_start_scaled[0]
                dy = gl_end_scaled[1] - gl_start_scaled[1]
                arrow_scale = 0.15
                ax.annotate('', xy=(gl_end_scaled[0], gl_end_scaled[1]),
                            xytext=(gl_end_scaled[0] - dx*arrow_scale, gl_end_scaled[1] - dy*arrow_scale),
                            arrowprops=dict(arrowstyle='->', color='black', lw=0.4))

    ax.set_xlim(-1, sheet_width_cm + 1)
    ax.set_ylim(-1, used_length_cm + 1)
    ax.set_aspect('equal')
    ax.set_xlabel('폭 (cm)')
    ax.set_ylabel('길이 (cm)')

    plt.tight_layout()
    return fig

# ==============================================================================
# 1. 페이지 및 스타일 설정 (Configuration & CSS)
# ==============================================================================
st.set_page_config(
    page_title="스마트 요척 산출서",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# 커스텀 CSS 스타일 정의
st.markdown("""
<style>
    /* 상단 여백 조정 - 제목이 잘리지 않도록 */
    .block-container {
        padding-top: 2rem;
        padding-bottom: 3rem;
        overflow: visible;
    }

    /* Streamlit 기본 헤더 숨기기 (여백 확보) */
    header[data-testid="stHeader"] {
        height: 0;
        min-height: 0;
        padding: 0;
        visibility: hidden;
    }

    /* 컴포넌트 간격 조정 */
    div[data-testid="stVerticalBlock"] > div { gap: 0rem; }
    div[data-testid="stColumn"] { text-align: center; }
    
    /* 숫자 버튼 스타일 (클릭 영역 확보) */
    div[data-testid="stColumn"] button {
        width: 100%;
        border: 1px solid #ccc;
        font-weight: bold;
        background-color: #f9f9f9;
        font-size: 13px;
        padding: 0px;
        margin-top: 2px;
        height: 28px;
    }
    div[data-testid="stColumn"] button:hover {
        border-color: #0068c9;
        color: #0068c9;
        background-color: #eef5ff;
    }
    
    /* 체크박스 중앙 정렬 보정 */
    div[data-testid="stColumn"] div[data-testid="stCheckbox"] {
        display: flex !important;
        justify-content: center !important;
        align-items: center !important;
        margin-top: -2px;
    }
    div[data-testid="stColumn"] div[data-testid="stCheckbox"] > label {
        justify-content: center !important;
    }
    div[data-testid="stColumn"] div[data-testid="stCheckbox"] > label > div {
        margin-right: 0 !important;
    }
    
    /* 기본 툴바 및 풀스크린 버튼 숨기기 (깔끔한 UI 유지) */
    [data-testid="stElementToolbar"] { display: none !important; }
    button[title="View fullscreen"] { display: none !important; }
    
    /* 요척 결과 카드 입력창 컴팩트 스타일 */
    div[data-testid="stNumberInput"] input {
        padding: 0px 5px;
        height: 30px;
        font-size: 13px;
        text-align: center;
    }

    /* 숫자 입력 플러스/마이너스 버튼 크기 (60%) */
    div[data-testid="stNumberInput"] button {
        width: 24px !important;
        min-width: 24px !important;
        padding: 0 6px !important;
    }
    div[data-testid="stNumberInput"] button svg {
        width: 12px !important;
        height: 12px !important;
    }

    /* 네스팅 결과 expander 제목 크기 25% 확대 */
    div[data-testid="stExpander"] summary span {
        font-size: 1.25em !important;
        font-weight: bold !important;
    }

    /* 네스팅 결과 메트릭 - 제목/데이터 모두 가운데 정렬 */
    div[data-testid="stExpander"] div[data-testid="stMetric"] {
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        text-align: center !important;
    }
    div[data-testid="stExpander"] div[data-testid="stMetric"] label {
        font-size: 0.7rem !important;
        width: 100% !important;
        text-align: center !important;
    }
    div[data-testid="stExpander"] div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        font-size: 1rem !important;
        width: 100% !important;
        text-align: center !important;
    }

    /* file_uploader 드래그앤드롭 영역 스타일 강화 */
    [data-testid="stFileUploader"],
    .stFileUploader {
        border: 3px dashed #0068c9 !important;
        border-radius: 15px !important;
        padding: 15px !important;
        background: linear-gradient(135deg, #f0f7ff 0%, #e8f4f8 100%) !important;
    }
    [data-testid="stFileUploader"] section,
    .stFileUploader section {
        min-height: 200px !important;
        display: flex !important;
        flex-direction: column !important;
        justify-content: center !important;
    }
    [data-testid="stFileUploader"] label,
    .stFileUploader label {
        font-size: 16px !important;
        font-weight: bold !important;
        color: #0068c9 !important;
    }
    [data-testid="stFileUploader"] small,
    .stFileUploader small {
        font-size: 13px !important;
    }

    /* 상세 리스트 데이터 에디터 - 헤더와 데이터 가운데 정렬 */
    div[data-testid="stDataFrame"] th,
    div[data-testid="stDataFrame"] td,
    div[data-testid="stDataEditor"] th,
    div[data-testid="stDataEditor"] td {
        text-align: center !important;
    }
    /* glide-data-grid 셀 가운데 정렬 */
    .dvn-scroller .dvn-cell,
    .gdg-cell {
        text-align: center !important;
        justify-content: center !important;
    }
    /* data_editor 헤더 가운데 정렬 */
    .gdg-header-cell,
    .dvn-header-cell {
        text-align: center !important;
        justify-content: center !important;
    }

    /* 요척 결과 메트릭 - 가운데 정렬 */
    div[data-testid="stMetric"] {
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        text-align: center !important;
    }
    div[data-testid="stMetric"] label {
        width: 100% !important;
        text-align: center !important;
    }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        width: 100% !important;
        text-align: center !important;
    }

    /* 네스팅 시뮬레이션 원단 설정 - 입력 필드 가운데 정렬 */
    div[data-testid="stNumberInput"] input {
        text-align: center !important;
    }
    div[data-testid="stTextInput"] input {
        text-align: center !important;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 style='font-size: 1.8rem; margin-top: 0; margin-bottom: 0.5rem; white-space: nowrap;'>👕 스마트 의류 요척(자동마카) 산출서</h1>", unsafe_allow_html=True)

# Streamlit 버전 호환성 체크 (팝업창 기능용)
try:
    st_version = st.__version__
    major, minor = map(int, st_version.split('.')[:2])
    if major < 1 or (major == 1 and minor < 34):
        st.error(f"🚨 중요: 현재 Streamlit 버전({st_version})이 낮습니다. 터미널에 'pip install --upgrade streamlit'을 실행해주세요.")
        st.stop()
except: pass


# ==============================================================================
# 2. 헬퍼 함수 및 유틸리티 (Helpers)
# ==============================================================================

def export_nesting_to_excel(nesting_results, timestamp, style_no=None, selected_sizes=None, base_size=None):
    """네스팅 결과를 엑셀로 내보내기 (한 시트에 모든 데이터 순서대로)

    Args:
        nesting_results: 네스팅 결과 딕셔너리
        timestamp: 작업일시
        style_no: 스타일번호 (DXF 파일명)
        selected_sizes: 선택된 사이즈 목록
        base_size: 기준사이즈
    """
    from io import BytesIO
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    import matplotlib.pyplot as plt

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "네스팅결과"

    # 스타일 정의
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    section_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=16)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1

    # === 0. 제목 정보 섹션 ===
    # 제목은 항상 표시
    ws.cell(row=current_row, column=1, value="📐 스마트 의류 요척(자동마카) 산출서").font = title_font
    current_row += 1

    # 스타일번호 표시
    if style_no and style_no.strip():
        ws.cell(row=current_row, column=1, value=f"스타일번호: {style_no}").font = header_font
        current_row += 1

    # 선택 사이즈 표시
    if selected_sizes and len(selected_sizes) > 0:
        size_display = ', '.join(selected_sizes)
        base_display = f" (기준: {base_size})" if base_size else ""
        ws.cell(row=current_row, column=1, value=f"선택 사이즈: {size_display}{base_display}").font = header_font
        current_row += 1

    current_row += 1  # 빈 줄 추가

    # === 1. 마카 요약 섹션 ===
    ws.cell(row=current_row, column=1, value="■ 마카 요약").font = section_font
    current_row += 1

    summary_headers = ['원단', '벌수', '패턴수', '원단폭(cm)', '마카길이(cm)', '요척(YD)', '효율(%)', '작업일시']
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    for fabric, result in nesting_results.items():
        if result.get('success'):
            # 벌수 계산: 사이즈별 벌수 합계 또는 단일 벌수
            size_quantities = result.get('size_quantities', {})
            has_multiple_sizes = result.get('has_multiple_sizes', False)
            if has_multiple_sizes and size_quantities and selected_sizes:
                marker_qty = sum(size_quantities.get(s, 1) for s in selected_sizes if size_quantities.get(s, 1) > 0)
            else:
                marker_qty = result.get('marker_quantity', 1)

            yield_per_set = result.get('used_length_yd', 0) / marker_qty if marker_qty > 0 else result.get('used_length_yd', 0)
            row_data = [
                fabric,
                marker_qty,
                f"{result.get('placed_count', 0)}/{result.get('total_count', 0)}",
                result.get('width_cm', 0),
                round(result.get('used_length_cm', 0), 1),
                round(yield_per_set, 2),
                result.get('efficiency', 0),
                timestamp
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

    current_row += 2  # 빈 줄 추가

    # === 2. 원단별 마카 이미지 + 배치 상세 (2열 배치) ===
    ws.cell(row=current_row, column=1, value="■ 마카 이미지").font = section_font
    current_row += 1

    fabric_list = [f for f, r in nesting_results.items() if r.get('success')]

    # 2개씩 묶어서 처리
    for i in range(0, len(fabric_list), 2):
        row_start = current_row
        img_rows1 = 0
        img_rows2 = 0

        # 왼쪽 마카 (A열)
        fabric1 = fabric_list[i]
        result1 = nesting_results[fabric1]
        ws.cell(row=current_row, column=1, value=f"▷ {fabric1}").font = Font(bold=True, size=11)

        try:
            width_cm = result1.get('width_cm', 150)
            if result1.get('sparrow_mode'):
                fig = create_sparrow_visualization(result1, width_cm)
            else:
                fig = create_nesting_visualization(result1, width_cm)

            if fig:
                img_buffer = BytesIO()
                fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight',
                           facecolor='white', edgecolor='none')
                img_buffer.seek(0)
                plt.close(fig)

                img = XLImage(img_buffer)
                orig_width = img.width
                orig_height = img.height
                if orig_width > 0:
                    img.width = 450
                    img.height = int(orig_height * (450 / orig_width))

                ws.add_image(img, f'A{current_row + 1}')
                img_rows1 = int(img.height / 15) + 2
        except Exception as e:
            ws.cell(row=current_row + 1, column=1, value=f"오류: {e}")
            img_rows1 = 3

        # 오른쪽 마카 (F열) - 있으면
        fabric2 = None
        result2 = None
        if i + 1 < len(fabric_list):
            fabric2 = fabric_list[i + 1]
            result2 = nesting_results[fabric2]
            ws.cell(row=current_row, column=6, value=f"▷ {fabric2}").font = Font(bold=True, size=11)

            try:
                width_cm = result2.get('width_cm', 150)
                if result2.get('sparrow_mode'):
                    fig = create_sparrow_visualization(result2, width_cm)
                else:
                    fig = create_nesting_visualization(result2, width_cm)

                if fig:
                    img_buffer = BytesIO()
                    fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight',
                               facecolor='white', edgecolor='none')
                    img_buffer.seek(0)
                    plt.close(fig)

                    img = XLImage(img_buffer)
                    orig_width = img.width
                    orig_height = img.height
                    if orig_width > 0:
                        img.width = 450
                        img.height = int(orig_height * (450 / orig_width))

                    ws.add_image(img, f'F{current_row + 1}')
                    img_rows2 = int(img.height / 15) + 2
            except Exception as e:
                ws.cell(row=current_row + 1, column=6, value=f"오류: {e}")
                img_rows2 = 3

        # 마카 이미지 아래로 이동
        max_img_rows = max(img_rows1, img_rows2, 10)
        current_row += max_img_rows + 2

    # 열 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 18

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_fabric_color_hex(fabric_name):
    """원단 이름에 따른 색상 코드를 반환합니다."""
    return get_fabric_color(fabric_name)

def extract_lines(entity, lines_list):
    """DXF 엔티티에서 선분 정보를 재귀적으로 추출합니다."""
    dxftype = entity.dxftype()
    try:
        if dxftype == 'LINE':
            start, end = entity.dxf.start, entity.dxf.end
            lines_list.append(LineString([(start.x, start.y), (end.x, end.y)]))
        elif dxftype in ['LWPOLYLINE', 'POLYLINE']:
            points = list(entity.points())
            if len(points) > 1:
                lines_list.append(LineString([(p[0], p[1]) for p in points]))
        elif dxftype in ['SPLINE', 'ARC', 'CIRCLE', 'ELLIPSE']:
            path = ezdxf.path.make_path(entity)
            vertices = list(path.flattening(distance=1.0))
            if len(vertices) > 1:
                lines_list.append(LineString([(v.x, v.y) for v in vertices]))
        elif dxftype == 'INSERT':
            # 블록 참조(Insert)일 경우 내부 엔티티 탐색
            for virtual_entity in entity.virtual_entities():
                extract_lines(virtual_entity, lines_list)
    except Exception:
        pass # 파싱 불가능한 엔티티는 무시

def check_is_fold(poly):
    """패턴이 골선(Fold)인지 판별합니다. (폭/높이가 매우 좁고 긴 형태)"""
    minx, miny, maxx, maxy = poly.bounds
    full_h, full_w = maxy - miny, maxx - minx
    coords = list(poly.exterior.coords)
    for i in range(len(coords)-1):
        p1, p2 = coords[i], coords[i+1]
        dist = math.sqrt((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2)
        # 수직골 또는 수평골 판별 로직
        # if abs(p1[0]-p2[0]) < 0.1 and dist > full_h * 0.9: return True, "수직골" (사용자 요청으로 삭제 - 대칭 로직으로 대체)
        # if abs(p1[1]-p2[1]) < 0.1 and dist > full_w * 0.9: return True, "수평골" (사용자 요청으로 삭제)
    return False, "일반"

def poly_to_base64(poly, fill_color='gray'):
    """Shapely Polygon을 정사각형 썸네일 이미지(Base64)로 변환합니다."""
    fig, ax = plt.subplots(figsize=(1, 1))
    x, y = poly.exterior.xy
    ax.plot(x, y, 'k-', lw=2)
    ax.fill(x, y, fill_color, alpha=0.6) # 색상 적용 (투명도 약간 높임)
    ax.axis('off')

    # 정사각형 비율 맞추기 (Centering)
    minx, miny, maxx, maxy = poly.bounds
    cx, cy = (minx + maxx) / 2, (miny + maxy) / 2
    max_dim = max(maxx - minx, maxy - miny)
    padding = max_dim * 0.1 # 10% 여백
    span = (max_dim + padding) / 2

    ax.set_xlim(cx - span, cx + span)
    ax.set_ylim(cy - span, cy + span)
    ax.set_aspect('equal')

    buf = io.BytesIO()
    fig.savefig(buf, format='png', transparent=True, bbox_inches='tight', pad_inches=0)
    plt.close(fig)

    data = base64.b64encode(buf.getvalue()).decode("utf-8")
    return f"data:image/png;base64,{data}"


def get_cached_thumbnail(idx, poly, fabric_name, zoom_span, grainline_info=None):
    """
    썸네일 캐싱 함수: (폴리곤 고유ID, 원단명, zoom_span) 조합으로 캐시 관리
    원단명 또는 zoom_span이 변경되면 해당 썸네일만 새로 생성
    복사/정렬 후에도 폴리곤 형상으로 정확히 매칭
    패턴 삭제 시 zoom_span 변경으로 자동 재생성
    그레인라인이 있으면 빨간 점선으로 표시
    """
    # 캐시 초기화
    if 'thumbnail_cache' not in st.session_state:
        st.session_state.thumbnail_cache = {}

    # 캐시 키: (폴리곤 면적 + 중심점 해시, 원단명, zoom_span, grainline 유무) - 폴리곤 형상 + 크기 기반 고유 식별
    poly_id = (round(poly.area, 2), round(poly.centroid.x, 2), round(poly.centroid.y, 2))
    zoom_key = round(zoom_span, 1)  # zoom_span 변경 감지
    has_grainline = grainline_info is not None
    cache_key = (poly_id, fabric_name, zoom_key, has_grainline)

    # 캐시에 있으면 재사용
    if cache_key in st.session_state.thumbnail_cache:
        return st.session_state.thumbnail_cache[cache_key]

    # 없으면 새로 생성
    fig, ax = plt.subplots(figsize=(1, 1))
    x, y = poly.exterior.xy
    ax.plot(x, y, 'k-', lw=0.5)
    ax.fill(x, y, color=get_fabric_color_hex(fabric_name), alpha=0.6)

    # 그레인라인 - 일괄수정도구 썸네일에서는 숨김 처리
    pass

    ax.set_xlim(poly.centroid.x - zoom_span/2, poly.centroid.x + zoom_span/2)
    ax.set_ylim(poly.centroid.y - zoom_span/2, poly.centroid.y + zoom_span/2)
    ax.set_aspect('equal')
    ax.axis('off')

    # BytesIO로 이미지 저장
    buf = io.BytesIO()
    fig.savefig(buf, format='png', transparent=True, bbox_inches='tight', pad_inches=0, dpi=72)
    plt.close(fig)
    buf.seek(0)

    # 캐시에 저장
    st.session_state.thumbnail_cache[cache_key] = buf.getvalue()

    return st.session_state.thumbnail_cache[cache_key]


def create_overlay_visualization(patterns_group, selected_sizes, all_sizes, global_max_dim=None, base_size=None):
    """
    동일 패턴 그룹의 여러 사이즈를 중첩하여 시각화 (Plotly 인터랙티브)
    - 바탕색 없이 외곽선만 표시
    - 사이즈별 다른 색상 외곽선
    - 중심 정렬로 크기 비교
    - 기준사이즈의 내부선(스티치선, 시접선) 표시
    - 마우스 휠로 확대/축소, 드래그로 패닝 가능

    Args:
        patterns_group: [(poly, pattern_name, fabric_name, size_name, pattern_group, ..., interior_lines), ...] 동일 그룹
        selected_sizes: 선택된 사이즈 목록
        all_sizes: 전체 사이즈 목록
        global_max_dim: 전역 최대 크기 (모든 패턴에 동일 비율 적용)
        base_size: 기준사이즈 (내부선 표시용)

    Returns:
        plotly figure
    """
    import plotly.graph_objects as go
    import matplotlib.pyplot as plt

    fig = go.Figure()

    # 사이즈별 색상 (파랑→빨강 그라데이션) - matplotlib colormap 활용
    size_colors = {}
    cmap = plt.cm.get_cmap('coolwarm', len(all_sizes) + 1)
    for i, size in enumerate(all_sizes):
        rgba = cmap(i / max(len(all_sizes) - 1, 1))
        # RGBA to hex
        size_colors[size] = f'rgba({int(rgba[0]*255)},{int(rgba[1]*255)},{int(rgba[2]*255)},{rgba[3]})'

    # 모든 패턴의 경계 계산 (중심 맞추기용)
    all_bounds = []
    for p_data in patterns_group:
        poly = p_data[0]
        all_bounds.append(poly.bounds)

    if not all_bounds:
        return fig

    # 전체 영역 계산
    min_x = min(b[0] for b in all_bounds)
    min_y = min(b[1] for b in all_bounds)
    max_x = max(b[2] for b in all_bounds)
    max_y = max(b[3] for b in all_bounds)

    center_x = (min_x + max_x) / 2
    center_y = (min_y + max_y) / 2

    # 사이즈 순서대로 그리기 (큰 것부터)
    sorted_patterns = sorted(patterns_group, key=lambda x: x[0].area, reverse=True)

    drawn_sizes = set()

    for p_data in sorted_patterns:
        poly = p_data[0]
        size_name = p_data[3]

        if not size_name:
            continue

        # 원본 DXF 위치 그대로 사용
        x, y = poly.exterior.xy
        x_list = list(x)
        y_list = list(y)

        # 사이즈별 색상
        color = size_colors.get(size_name, 'rgba(128,128,128,1)')
        is_selected = size_name in selected_sizes

        # 선택된 사이즈: 실선, 미선택: 점선+흐리게
        show_legend = size_name not in drawn_sizes
        drawn_sizes.add(size_name)

        if is_selected:
            fig.add_trace(go.Scatter(
                x=x_list, y=y_list,
                mode='lines',
                line=dict(color=color, width=1.5),
                name=size_name,
                showlegend=show_legend,
                hoverinfo='name'
            ))
        else:
            fig.add_trace(go.Scatter(
                x=x_list, y=y_list,
                mode='lines',
                line=dict(color=color, width=0.5, dash='dash'),
                opacity=0.4,
                name=size_name,
                showlegend=show_legend,
                hoverinfo='name'
            ))

        # 기준사이즈인 경우 내부선(스티치선, 시접선) 표시
        if base_size and size_name == base_size and is_selected:
            if len(p_data) > 8 and p_data[8]:
                interior_lines = p_data[8]
                for line_coords in interior_lines:
                    if len(line_coords) >= 2:
                        line_x = [c[0] for c in line_coords]
                        line_y = [c[1] for c in line_coords]
                        fig.add_trace(go.Scatter(
                            x=line_x, y=line_y,
                            mode='lines',
                            line=dict(color=color, width=1.5),
                            opacity=1.0,
                            showlegend=False,
                            hoverinfo='skip'
                        ))

    # 축 설정 - 전역 최대 크기 사용 (모든 패턴 동일 비율)
    if global_max_dim:
        margin = global_max_dim * 0.15
        half_dim = global_max_dim / 2 + margin
        x_range = [center_x - half_dim, center_x + half_dim]
        y_range = [center_y - half_dim, center_y + half_dim]
    else:
        margin = max(max_x - min_x, max_y - min_y) * 0.15
        x_range = [min_x - margin, max_x + margin]
        y_range = [min_y - margin, max_y + margin]

    fig.update_layout(
        xaxis=dict(
            range=x_range,
            scaleanchor='y',
            scaleratio=1,
            showgrid=False,
            showticklabels=False,
            zeroline=False
        ),
        yaxis=dict(
            range=y_range,
            showgrid=False,
            showticklabels=False,
            zeroline=False
        ),
        showlegend=True,
        legend=dict(
            orientation='h',
            yanchor='bottom',
            y=1.02,
            xanchor='center',
            x=0.5,
            font=dict(size=10)
        ),
        margin=dict(l=5, r=5, t=30, b=5),
        height=300,
        dragmode='pan',  # 기본 드래그 모드를 팬으로 설정
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )

    # 확대/축소, 패닝 활성화
    fig.update_xaxes(fixedrange=False)
    fig.update_yaxes(fixedrange=False)

    return fig


def check_symmetry(poly):
    """
    패턴의 대칭 여부를 판단합니다. (좌우대칭 or 상하대칭)
    원리: 중심축 기준으로 반전시켰을 때 원본과 거의 겹치는지(차집합 면적이 적은지) 확인
    """
    try:
        # 허용 오차 (전체 면적의 2% 미만 차이면 대칭으로 간주 - 98% 일치)
        tolerance = poly.area * 0.02 
        
        # 1. 좌우 대칭 확인 (Horizontal Reflection)
        reflected_h = affinity.scale(poly, xfact=-1, origin='centroid')
        diff_h = poly.symmetric_difference(reflected_h).area
        if diff_h < tolerance:
            return True, "좌우대칭"

        # 2. 상하 대칭 확인 (Vertical Reflection)
        reflected_v = affinity.scale(poly, yfact=-1, origin='centroid')
        diff_v = poly.symmetric_difference(reflected_v).area
        if diff_v < tolerance:
            return True, "상하대칭"
            
        return False, "비대칭"
    except:
        return False, "오류"


def check_vertical_straight_edge(poly, min_length_cm=50):
    """
    패턴의 세로 직선(좌측/우측)이 특정 길이 이상인지 판별합니다.
    min_length_cm 이상인 세로 직선이 1개라도 있으면 True 반환
    """
    try:
        coords = list(poly.exterior.coords)
        if len(coords) < 4:
            return False

        min_length_mm = min_length_cm * 10  # cm → mm 변환

        # 연속된 점들 사이의 세로 직선 확인
        for i in range(len(coords) - 1):
            x1, y1 = coords[i]
            x2, y2 = coords[i + 1]

            # X좌표 변화가 거의 없으면 세로 직선
            x_diff = abs(x2 - x1)
            y_length = abs(y2 - y1)

            # X좌표 변화가 전체 폭의 2% 이내이고, 세로 길이가 min_length 이상
            minx, miny, maxx, maxy = poly.bounds
            width = maxx - minx

            if x_diff < width * 0.02 and y_length >= min_length_mm:
                return True

        return False
    except:
        return False


def check_horizontal_straight_edge(poly):
    """
    패턴의 가로 직선(상단/하단)이 1개 이상 있는지 판별합니다.
    상단 또는 하단 중 Y좌표 변화가 1% 이내인 직선이 있으면 True 반환
    """
    try:
        coords = list(poly.exterior.coords)
        if len(coords) < 4:
            return False

        minx, miny, maxx, maxy = poly.bounds
        height = maxy - miny

        # 상단/하단 영역 정의 (전체 높이의 10% 이내)
        top_threshold = maxy - height * 0.1
        bottom_threshold = miny + height * 0.1

        top_points = [(x, y) for x, y in coords if y >= top_threshold]
        bottom_points = [(x, y) for x, y in coords if y <= bottom_threshold]

        def is_straight_line(points):
            if len(points) < 2:
                return False
            y_values = [p[1] for p in points]
            y_range = max(y_values) - min(y_values)
            return y_range < height * 0.01

        return is_straight_line(top_points) or is_straight_line(bottom_points)
    except:
        return False


def check_parallel_edges(poly, similarity_threshold=0.85):
    """
    패턴의 상하단 가로선이 평행선인지 판별합니다.
    상하단 가로 길이 비율이 similarity_threshold 이상이면 True 반환
    """
    try:
        coords = list(poly.exterior.coords)
        if len(coords) < 4:
            return False

        minx, miny, maxx, maxy = poly.bounds
        height = maxy - miny

        # 상단/하단 영역 정의 (전체 높이의 10% 이내)
        top_threshold = maxy - height * 0.1
        bottom_threshold = miny + height * 0.1

        top_points = [(x, y) for x, y in coords if y >= top_threshold]
        bottom_points = [(x, y) for x, y in coords if y <= bottom_threshold]

        def get_edge_length(points):
            if len(points) < 2:
                return 0
            x_values = [p[0] for p in points]
            return max(x_values) - min(x_values)

        top_length = get_edge_length(top_points)
        bottom_length = get_edge_length(bottom_points)

        if top_length > 0 and bottom_length > 0:
            similarity = min(top_length, bottom_length) / max(top_length, bottom_length)
            return similarity >= similarity_threshold

        return False
    except:
        return False


# ==============================================================================
# 3. 핵심 로직: DXF 처리 (Core Logic)
# ==============================================================================

from shapely.ops import polygonize, linemerge

# ... (기존 extract_lines 함수는 그대로 유지하거나 필요시 수정) ...

@st.cache_data
def extract_style_no(file_path):
    """DXF 파일에서 스타일번호를 추출합니다."""
    try:
        try:
            doc = ezdxf.readfile(file_path, encoding='cp949')
        except:
            doc = ezdxf.readfile(file_path)

        msp = doc.modelspace()

        for entity in msp:
            if entity.dxftype() == 'INSERT':
                block = doc.blocks.get(entity.dxf.name)
                for be in block:
                    if be.dxftype() == 'TEXT':
                        text = be.dxf.text
                        text_upper = text.upper()
                        # 스타일번호: S/#..., M/#... 형식 (대소문자 무시)
                        if text_upper.startswith('ANNOTATION:') and '/#' in text:
                            val = text.split(':', 1)[1].strip()
                            # S/#5535-731 → 5535-731
                            if '/#' in val:
                                return val.split('/#')[1]
                break  # 첫 번째 블록에서만 추출
    except:
        pass
    return ""


def is_symmetric_polygon(poly, tolerance=0.02):
    """
    폴리곤이 좌우대칭인지 판단합니다.

    Args:
        poly: Shapely Polygon
        tolerance: 대칭 판단 허용 오차 (면적 비율, 기본 2%)

    Returns:
        bool: 좌우대칭이면 True
    """
    from shapely import affinity
    from shapely.ops import unary_union

    if not poly or poly.is_empty:
        return False

    # 폴리곤의 중심과 바운딩 박스
    centroid = poly.centroid
    minx, miny, maxx, maxy = poly.bounds

    # Y축 (수직 중심선) 기준으로 좌우 반전
    mirrored = affinity.scale(poly, xfact=-1, yfact=1, origin=centroid)

    # 원본과 반전본의 교집합 면적 비교
    try:
        intersection = poly.intersection(mirrored)
        intersection_area = intersection.area
        original_area = poly.area

        # 교집합 면적이 원본의 (1-tolerance) 이상이면 대칭
        if original_area > 0 and intersection_area / original_area >= (1 - tolerance):
            return True
    except:
        pass

    return False


def get_center_vertical_grainline(poly):
    """
    폴리곤의 중심 수직선 좌표를 반환합니다.

    Args:
        poly: Shapely Polygon

    Returns:
        (angle, start, end): 각도 90도, 시작점, 끝점
    """
    minx, miny, maxx, maxy = poly.bounds
    center_x = (minx + maxx) / 2
    height = maxy - miny

    # 패턴 높이의 70% 정도 길이로 결선 생성
    line_length = height * 0.7
    center_y = (miny + maxy) / 2

    start = (center_x, center_y + line_length / 2)
    end = (center_x, center_y - line_length / 2)

    return 90.0, start, end


def detect_grainline(block, return_coords=False, polygon=None):
    """
    블록 내에서 그레인라인(결방향)을 감지합니다.

    그레인라인 감지 방법:
    1. 레이어명에 'GRAIN', 'GL', '결', '7' 등 포함된 엔티티
    2. 그레인라인 레이어가 없으면 독립 LINE 중 가장 긴 것 (100단위 이상)
    3. 위 방법으로도 없으면 좌우대칭 패턴인 경우 중심 수직선 사용

    Args:
        block: ezdxf 블록 객체
        return_coords: True면 (angle, start, end) 반환, False면 angle만 반환
        polygon: Shapely Polygon (좌우대칭 판단용, 선택)

    Returns:
        return_coords=False: angle (그레인라인 각도) 또는 None
        return_coords=True: (angle, start, end) 또는 (None, None, None)
    """
    import math

    candidate_lines = []

    for entity in block:
        layer_name = str(entity.dxf.layer).upper() if hasattr(entity.dxf, 'layer') else ''

        # 레이어명으로 그레인라인 감지 (키워드 또는 숫자 레이어)
        is_grainline_layer = (
            any(kw in layer_name for kw in GRAINLINE_KEYWORDS) or
            layer_name in GRAINLINE_LAYER_NUMBERS
        )

        if entity.dxftype() == 'LINE':
            start = entity.dxf.start
            end = entity.dxf.end
            dx = end.x - start.x
            dy = end.y - start.y
            length = math.sqrt(dx*dx + dy*dy)

            if length > 1:  # 최소 길이 필터
                # 각도 계산 (수평 기준, -180 ~ 180)
                angle = math.degrees(math.atan2(dy, dx))
                candidate_lines.append({
                    'angle': angle,
                    'length': length,
                    'is_grainline_layer': is_grainline_layer,
                    'start': (start.x, start.y),
                    'end': (end.x, end.y)
                })

        elif entity.dxftype() == 'LWPOLYLINE' and not entity.closed:
            # 열린 폴리라인도 그레인라인일 수 있음
            pts = list(entity.get_points())
            if len(pts) == 2:  # 2점 직선
                start, end = pts[0], pts[1]
                dx = end[0] - start[0]
                dy = end[1] - start[1]
                length = math.sqrt(dx*dx + dy*dy)
                if length > 1:
                    angle = math.degrees(math.atan2(dy, dx))
                    candidate_lines.append({
                        'angle': angle,
                        'length': length,
                        'is_grainline_layer': is_grainline_layer,
                        'start': (start[0], start[1]),
                        'end': (end[0], end[1])
                    })

    if not candidate_lines:
        # 후보 라인이 없으면 좌우대칭 체크
        if polygon and is_symmetric_polygon(polygon):
            return get_center_vertical_grainline(polygon) if return_coords else 90.0
        return (None, None, None) if return_coords else None

    # 그레인라인 우선순위:
    # 1. 그레인라인 레이어에 있는 선
    # 2. 그 중 가장 긴 선
    grainline_layer_lines = [l for l in candidate_lines if l['is_grainline_layer']]

    best = None
    if grainline_layer_lines:
        # 그레인라인 레이어에서 가장 긴 선
        best = max(grainline_layer_lines, key=lambda x: x['length'])
    else:
        # 그레인라인 레이어가 없으면 독립 LINE 중 가장 긴 것 사용 (100단위 이상)
        long_lines = [l for l in candidate_lines if l['length'] >= 100]
        if long_lines:
            best = max(long_lines, key=lambda x: x['length'])

    # 좌우대칭 패턴인 경우: 결선이 중심에서 벗어났으면 중심 수직선 사용
    if best and polygon and is_symmetric_polygon(polygon):
        minx, miny, maxx, maxy = polygon.bounds
        center_x = (minx + maxx) / 2
        width = maxx - minx

        # 결선의 중심 X 좌표
        gl_center_x = (best['start'][0] + best['end'][0]) / 2

        # 결선이 패턴 중심에서 폭의 5% 이상 벗어났으면 중심 수직선 사용
        if abs(gl_center_x - center_x) > width * 0.05:
            return get_center_vertical_grainline(polygon) if return_coords else 90.0

    if best:
        if return_coords:
            return best['angle'], best['start'], best['end']
        return best['angle']

    # 최종적으로 결선을 찾지 못하면 좌우대칭 체크
    if polygon and is_symmetric_polygon(polygon):
        return get_center_vertical_grainline(polygon) if return_coords else 90.0

    return (None, None, None) if return_coords else None


def rotate_polygon_to_vertical_grain(poly, grainline_angle):
    """
    폴리곤을 그레인라인이 수직(90도)이 되도록 회전합니다.

    Args:
        poly: Shapely Polygon
        grainline_angle: 현재 그레인라인 각도 (도)

    Returns:
        회전된 Shapely Polygon
    """
    from shapely import affinity

    # 수직(90도)으로 맞추기 위해 필요한 회전 각도
    # grainline_angle이 0도면 90도 회전 필요
    # grainline_angle이 45도면 45도 회전 필요
    # grainline_angle이 90도면 0도 회전 (이미 수직)

    # 각도 정규화 (-90 ~ 90 범위로)
    normalized = grainline_angle % 180
    if normalized > 90:
        normalized -= 180

    # 수직(90도)까지 회전 필요 각도
    rotation_needed = 90 - normalized

    # 180도 이상 회전 방지
    if rotation_needed > 90:
        rotation_needed -= 180
    elif rotation_needed < -90:
        rotation_needed += 180

    if abs(rotation_needed) < 1:  # 이미 수직에 가까움
        return poly, 0

    # 중심점 기준 회전
    rotated = affinity.rotate(poly, rotation_needed, origin='centroid')
    return rotated, rotation_needed


def rotate_grainline_coords(start, end, rotation_angle, origin):
    """
    그레인라인 좌표를 회전시킵니다.

    Args:
        start: (x, y) 시작점
        end: (x, y) 끝점
        rotation_angle: 회전 각도 (도)
        origin: (x, y) 회전 중심점

    Returns:
        (new_start, new_end) 회전된 좌표
    """
    import math

    if abs(rotation_angle) < 1:
        return start, end

    rad = math.radians(rotation_angle)
    cos_a = math.cos(rad)
    sin_a = math.sin(rad)
    ox, oy = origin

    def rotate_point(px, py):
        dx = px - ox
        dy = py - oy
        new_x = ox + dx * cos_a - dy * sin_a
        new_y = oy + dx * sin_a + dy * cos_a
        return (new_x, new_y)

    new_start = rotate_point(start[0], start[1])
    new_end = rotate_point(end[0], end[1])
    return new_start, new_end


def detect_grainline_for_polygon(msp, poly):
    """
    모델스페이스에서 특정 폴리곤 내부 또는 근처에 있는 그레인라인을 감지합니다.
    레거시 DXF 파일용 (블록 없이 모델스페이스에 직접 그려진 패턴)

    Args:
        msp: ezdxf 모델스페이스
        poly: Shapely Polygon (대상 패턴)

    Returns:
        angle: 그레인라인의 각도 (도)
        None: 그레인라인을 찾지 못한 경우
    """
    import math
    from shapely.geometry import Point, LineString

    candidate_lines = []

    # 폴리곤 바운딩박스 확장 (근처 선 검색용)
    bounds = poly.bounds
    margin = 50  # 50단위 여유
    search_bounds = (bounds[0] - margin, bounds[1] - margin,
                     bounds[2] + margin, bounds[3] + margin)

    for entity in msp:
        layer_name = entity.dxf.layer.upper() if hasattr(entity.dxf, 'layer') else ''
        is_grainline_layer = any(kw in layer_name for kw in GRAINLINE_KEYWORDS)

        if entity.dxftype() == 'LINE':
            start = entity.dxf.start
            end = entity.dxf.end

            # 바운딩박스 내에 있는지 확인
            mid_x = (start.x + end.x) / 2
            mid_y = (start.y + end.y) / 2
            if not (search_bounds[0] <= mid_x <= search_bounds[2] and
                    search_bounds[1] <= mid_y <= search_bounds[3]):
                continue

            # 폴리곤 내부 또는 근처에 있는지 확인
            mid_point = Point(mid_x, mid_y)
            if not (poly.contains(mid_point) or poly.distance(mid_point) < margin):
                continue

            dx = end.x - start.x
            dy = end.y - start.y
            length = math.sqrt(dx*dx + dy*dy)

            if length > 1:
                angle = math.degrees(math.atan2(dy, dx))
                candidate_lines.append({
                    'angle': angle,
                    'length': length,
                    'is_grainline_layer': is_grainline_layer
                })

    if not candidate_lines:
        # 후보 라인이 없으면 좌우대칭 체크
        if is_symmetric_polygon(poly):
            return 90.0  # 좌우대칭이면 수직(90도) 결선
        return None

    # 그레인라인 레이어에 있는 선 우선
    grainline_layer_lines = [l for l in candidate_lines if l['is_grainline_layer']]
    if grainline_layer_lines:
        best = max(grainline_layer_lines, key=lambda x: x['length'])
        return best['angle']

    # 최종적으로 결선을 찾지 못하면 좌우대칭 체크
    if is_symmetric_polygon(poly):
        return 90.0

    return None


def preprocess_dxf_content(file_path):
    """
    DXF 파일 내용을 전처리하여 특수문자 문제를 해결합니다.
    블록명에 <&> 등 ezdxf가 처리할 수 없는 문자가 있으면 치환합니다.
    """
    import tempfile
    import os

    # 파일 읽기 (CP949 우선)
    content = None
    for encoding in ['cp949', 'utf-8', 'latin-1']:
        try:
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                content = f.read()
            break
        except:
            continue

    if content is None:
        return file_path  # 읽기 실패 시 원본 반환

    # 특수문자 치환이 필요한지 확인
    if '<&>' not in content and '<>' not in content:
        return file_path  # 치환 불필요

    # 특수문자 치환 (블록명에서 문제가 되는 문자들)
    content = content.replace('<&>', 'X')  # <&> → X로 치환
    content = content.replace('<>', 'XX')  # <> → XX로 치환

    # 임시 파일로 저장
    temp_dir = tempfile.gettempdir()
    temp_filename = os.path.join(temp_dir, f'dxf_preprocessed_{os.path.basename(file_path)}')
    with open(temp_filename, 'w', encoding='cp949', errors='ignore') as f:
        f.write(content)

    return temp_filename


def scan_dxf_sizes(file_path):
    """
    DXF 파일에서 사이즈 목록만 빠르게 스캔합니다.
    전체 파싱 없이 사이즈 정보만 추출하여 선택 UI에 사용합니다.

    Returns:
        tuple: (사이즈 목록, 기준사이즈) - 기준사이즈가 없으면 중간 사이즈
    """
    import re

    sizes = set()
    base_size = None  # 기준사이즈

    try:
        # 특수문자 전처리
        processed_path = preprocess_dxf_content(file_path)

        # 한글 인코딩(CP949) 우선 시도
        try:
            doc = ezdxf.readfile(processed_path, encoding='cp949')
        except:
            doc = ezdxf.readfile(processed_path)

        msp = doc.modelspace()

        # 방법 0: modelspace 직접 TEXT에서 기준사이즈 추출
        for entity in msp:
            if entity.dxftype() == 'TEXT' and not base_size:
                text = entity.dxf.text
                text_upper = text.upper().strip()
                for prefix in BASE_SIZE_PREFIXES:
                    if text_upper.startswith(prefix):
                        base_val = text.split(':', 1)[1].strip().upper()
                        if base_val:
                            # 영문/숫자만 추출 (예: "M축적용" → "M")
                            base_size = extract_english_size(base_val)
                        break

        # 방법 1: INSERT 블록에서 사이즈 추출
        for entity in msp:
            if entity.dxftype() == 'INSERT':
                block_name = entity.dxf.name

                # 블록명에서 사이즈 추출 (예: BLK_1_XS, 앞판_M)
                if '_' in block_name:
                    _, potential_size = block_name.rsplit('_', 1)
                    if re.match(r'^([0-9]*X{1,3}L?|[SML]|XS|\d{2,3})$', potential_size, re.IGNORECASE):
                        sizes.add(potential_size.upper())

                # 블록 내 TEXT에서 SIZE: 및 기준사이즈 필드 추출
                try:
                    block = doc.blocks.get(block_name)
                    for be in block:
                        if be.dxftype() == 'TEXT':
                            text = be.dxf.text
                            text_upper = text.upper().strip()
                            if text_upper.startswith('SIZE:'):
                                size_val = text.split(':', 1)[1].strip()
                                if size_val:
                                    # 영문/숫자만 추출 (예: "S축적용" → "S")
                                    size_val = extract_english_size(size_val)
                                    sizes.add(size_val.upper())
                            # 기준사이즈 추출 (여러 패턴 지원)
                            elif not base_size:
                                for prefix in BASE_SIZE_PREFIXES:
                                    if text_upper.startswith(prefix):
                                        base_val = text.split(':', 1)[1].strip().upper()
                                        if base_val:
                                            # 영문/숫자만 추출
                                            base_size = extract_english_size(base_val)
                                        break
                except:
                    pass
    except Exception as e:
        st.error(f"사이즈 스캔 오류: {e}")
        return [], None

    # 사이즈 정렬 (constants.size_sort_key 사용)
    sorted_sizes = sorted(list(sizes), key=size_sort_key)

    # 기준사이즈가 없으면 중간 사이즈를 기준으로 설정
    if not base_size and sorted_sizes:
        mid_idx = len(sorted_sizes) // 2
        base_size = sorted_sizes[mid_idx]

    return sorted_sizes, base_size


@st.cache_data
def process_dxf(file_path, selected_sizes=None):
    """
    DXF 파일을 읽어 (Polygon, 패턴이름, 원단명, 사이즈) 튜플 리스트를 반환합니다.
    블록(INSERT) 기반으로 처리하여 패턴 누락을 방지합니다.

    Args:
        file_path: DXF 파일 경로
        selected_sizes: 선택된 사이즈 목록 (None이면 전체 로딩)
    """
    import re

    try:
        # 특수문자 전처리 (블록명에 <&> 등이 있으면 치환)
        processed_path = preprocess_dxf_content(file_path)

        # 한글 인코딩(CP949) 우선 시도
        try:
            doc = ezdxf.readfile(processed_path, encoding='cp949')
        except:
            doc = ezdxf.readfile(processed_path)

        msp = doc.modelspace()

        final = []
        detected_base_size = None  # DXF에서 추출한 기준사이즈
        unit_scale = 1.0  # 단위 스케일 (인치→mm 변환용)

        # 단위 정보 검색 (Units: ENGLISH → 인치, Units: METRIC → cm/mm)
        # 시스템 내부 단위는 mm이므로 인치→mm로 변환
        for entity in msp:
            if entity.dxftype() == 'TEXT':
                text = entity.dxf.text.upper().strip()
                if text.startswith('UNITS:'):
                    unit_value = text.split(':', 1)[1].strip()
                    if unit_value == 'ENGLISH':
                        unit_scale = UNIT_SCALE_INCH_TO_MM
                    elif unit_value == 'METRIC':
                        unit_scale = 1.0  # 이미 mm
                    break

        # 모델스페이스에서 기준사이즈 먼저 검색 (SAMPLE SIZE, BASE SIZE 등)
        for entity in msp:
            if entity.dxftype() == 'TEXT':
                text = entity.dxf.text
                text_upper = text.upper().strip()
                for prefix in BASE_SIZE_PREFIXES:
                    if text_upper.startswith(prefix):
                        base_val = text.split(':', 1)[1].strip()
                        if base_val and not detected_base_size:
                            detected_base_size = base_val
                        break
            if detected_base_size:
                break

        # 방법 1: INSERT 블록 기반 추출 (YUKA CAD 등)
        insert_count = 0
        for entity in msp:
            if entity.dxftype() == 'INSERT':
                insert_count += 1
                block_name = entity.dxf.name
                try:
                    block = doc.blocks.get(block_name)
                    max_poly = None
                    max_area = 0
                    pattern_name = ""
                    fabric_name = ""  # 원단명 추출용
                    size_name = ""    # 사이즈 추출용
                    pattern_group = ""  # 패턴 그룹 번호 (블록명에서 추출)
                    piece_name = ""   # PIECE NAME 필드
                    dxf_quantity = 0  # QUANTITY 필드 (원본 수량)

                    # 블록명에서 패턴그룹/사이즈 추출
                    # 형식1: BLK_1_XS → 그룹:1, 사이즈:XS (AAMA)
                    # 형식2: 앞판-a_M → 그룹:앞판-a, 사이즈:M
                    # 형식3: 1, 2, 3... → 그룹:블록명 (TIIP - 사이즈는 SIZE 필드에서)
                    if '_' in block_name:
                        base_name, potential_size = block_name.rsplit('_', 1)
                        # 사이즈 패턴: S, M, L, XS, XL, XXL, 2XL, 3XL, 0X, 1X, 00X, 85, 90 등
                        if re.match(r'^([0-9]*X{1,2}L?|[SML]|XS|\d{2,3})$', potential_size, re.IGNORECASE):
                            size_name = potential_size
                            pattern_group = base_name  # 사이즈 앞부분 전체를 그룹으로 사용
                        else:
                            # 사이즈 패턴이 아니면 기존 방식 시도 (BLK_1_0X 형식)
                            block_parts = block_name.split('_')
                            if len(block_parts) >= 2 and block_parts[1].isdigit():
                                pattern_group = block_parts[1]
                    else:
                        # TIIP 형식: 블록명이 숫자인 경우 (1, 2, 3...)
                        # pattern_group은 나중에 piece_name으로 설정됨
                        pass

                    # 블록 내 가장 큰 닫힌 POLYLINE 선택 + 텍스트 추출
                    for be in block:
                        if be.dxftype() == 'POLYLINE' and be.is_closed:
                            pts = list(be.points())
                            if len(pts) >= 3:
                                coords = [(p[0], p[1]) for p in pts]
                                poly = Polygon(coords)
                                # 유효하지 않은 폴리곤은 buffer(0)으로 수정 시도
                                if not poly.is_valid:
                                    poly = poly.buffer(0)
                                if poly.is_valid and poly.area > max_area:
                                    max_area = poly.area
                                    max_poly = poly
                        elif be.dxftype() == 'LWPOLYLINE' and be.closed:
                            pts = list(be.points())
                            if len(pts) >= 3:
                                coords = [(p[0], p[1]) for p in pts]
                                poly = Polygon(coords)
                                # 유효하지 않은 폴리곤은 buffer(0)으로 수정 시도
                                if not poly.is_valid:
                                    poly = poly.buffer(0)
                                if poly.is_valid and poly.area > max_area:
                                    max_area = poly.area
                                    max_poly = poly
                        elif be.dxftype() == 'TEXT':
                            text = be.dxf.text

                            # PIECE NAME / PIECE 필드에서 패턴 번호/이름 추출 (대소문자 무시)
                            text_upper = text.upper()
                            if text_upper.startswith('PIECE NAME:'):
                                piece_val = text.split(':', 1)[1].strip()
                                if piece_val:
                                    piece_name = piece_val
                            # TIIP 형식: PIECE: (PIECE NAME: 없이)
                            elif text_upper.startswith('PIECE:') and not text_upper.startswith('PIECE NAME:'):
                                piece_val = text.split(':', 1)[1].strip()
                                if piece_val:
                                    piece_name = piece_val
                                    # TIIP에서는 PIECE가 패턴명 역할
                                    if not pattern_name:
                                        pattern_name = piece_val

                            # QUANTITY 필드에서 원본 수량 추출 (대소문자 무시)
                            elif text_upper.startswith('QUANTITY:') or text_upper.startswith('QTY:'):
                                qty_val = text.split(':', 1)[1].strip()
                                if qty_val and qty_val.isdigit():
                                    dxf_quantity = int(qty_val)

                            # SIZE 필드에서 사이즈 추출 (대소문자 무시)
                            elif text_upper.startswith('SIZE:'):
                                size_val = text.split(':', 1)[1].strip()
                                if size_val:
                                    # 영문/숫자만 추출 (예: "S축적용" → "S")
                                    size_name = extract_english_size(size_val)

                            # CATEGORY 필드에서 원단명 추출 (대소문자 무시)
                            elif text_upper.startswith('CATEGORY:'):
                                cat_val = text.split(':', 1)[1].strip()
                                if cat_val:
                                    # 매핑된 원단명 찾기
                                    for key, mapped in FABRIC_MAP.items():
                                        if key.upper() == cat_val.upper() or key == cat_val:
                                            fabric_name = mapped
                                            break
                                    # 매핑 안 되면 기본값 "겉감" 사용
                                    if not fabric_name:
                                        fabric_name = "겉감"

                            # TIIP 형식: FABRIC 필드에서 원단명 추출
                            elif text_upper.startswith('FABRIC:'):
                                fab_val = text.split(':', 1)[1].strip()
                                if fab_val:
                                    # 매핑된 원단명 찾기
                                    for key, mapped in FABRIC_MAP.items():
                                        if key.upper() == fab_val.upper() or key == fab_val:
                                            fabric_name = mapped
                                            break
                                    # 매핑 안 되면 원본값 사용 (빈 문자열이면 기본값)
                                    if not fabric_name and fab_val:
                                        fabric_name = fab_val

                            # ANNOTATION 필드 처리 (대소문자 무시)
                            elif text_upper.startswith('ANNOTATION:'):
                                val = text.split(':', 1)[1].strip()
                                if not val:
                                    continue

                                # ANNOTATION에서 원단명 키워드 체크 (LINING 등)
                                val_upper = val.upper()
                                if val_upper in FABRIC_MAP:
                                    if not fabric_name:  # CATEGORY가 없을 때만
                                        fabric_name = FABRIC_MAP[val_upper]
                                    continue

                                # 사이즈 호칭 추출: <S>, <M>, <L>, <0X> 등
                                if val.startswith('<') and val.endswith('>'):
                                    extracted_size = val[1:-1].strip()
                                    if extracted_size and not size_name:
                                        # 영문/숫자만 추출 (예: "S축적용" → "S")
                                        size_name = extract_english_size(extracted_size)
                                    continue

                                # 제외 대상 체크
                                # 스타일 번호: S/#..., M/#... 등
                                if val.startswith(('S/', 'M/', 'L/', '#')):
                                    continue
                                # 숫자만 (사이즈: 130, 80 등)
                                if val.isdigit():
                                    continue
                                # 숫자로 시작 (스타일명: 35717요척 등)
                                if val[0].isdigit():
                                    continue
                                # 원단명 (이미 위에서 처리됨)
                                fabric_keywords = ['LINING', 'SHELL', 'INTERLINING', '안감', '겉감', '심지']
                                if val.upper() in [f.upper() for f in fabric_keywords]:
                                    continue
                                # 배색 관련
                                if '배색' in val:
                                    continue
                                # 문장 제외 (공백 2개 이상 또는 길이 5자 초과)
                                if val.count(' ') >= 2 or len(val) > 5:
                                    continue
                                # 괄호가 있는 설명문 제외
                                if '(' in val or ')' in val:
                                    continue
                                # 한글 부위명 우선 (한글이 포함되면 우선 선택)
                                has_korean = any('\uac00' <= c <= '\ud7a3' for c in val)
                                if has_korean:
                                    pattern_name = val  # 한글 부위명 덮어쓰기
                                elif not pattern_name:
                                    pattern_name = val  # 영문 부위명 (한글 없을 때만)

                            # TIIP 형식: COMMENT 필드 처리 (ANNOTATION과 유사)
                            elif text_upper.startswith('COMMENT:'):
                                val = text.split(':', 1)[1].strip()
                                if not val:
                                    continue
                                # COMMENT에서 원단명 키워드 체크
                                val_upper = val.upper()
                                if val_upper in FABRIC_MAP:
                                    if not fabric_name:
                                        fabric_name = FABRIC_MAP[val_upper]
                                    continue
                                # 제외 대상: 숫자만, cm/mm 포함, 너무 긴 문자열
                                if val.isdigit():
                                    continue
                                if 'cm' in val.lower() or 'mm' in val.lower():
                                    continue
                                if len(val) > 15:  # COMMENT는 설명이 길 수 있음
                                    continue
                                # 괄호 안 내용 제거 후 패턴명으로 사용
                                clean_val = re.sub(r'\([^)]*\)', '', val).strip()
                                if clean_val and not pattern_name:
                                    # 한글이 포함된 짧은 이름만 패턴명으로
                                    has_korean = any('\uac00' <= c <= '\ud7a3' for c in clean_val)
                                    if has_korean and len(clean_val) <= 10:
                                        pattern_name = clean_val

                    # 닫힌 POLYLINE/LWPOLYLINE이 없으면 열린 선분들을 연결하여 폴리곤 생성
                    if not max_poly:
                        block_lines = []
                        seen_lines = set()  # 중복 선분 제거용

                        for be in block:
                            if be.dxftype() == 'POLYLINE' and not be.is_closed:
                                pts = list(be.points())
                                if len(pts) >= 2:
                                    coords = tuple((round(p[0], 4), round(p[1], 4)) for p in pts)
                                    coords_rev = tuple(reversed(coords))
                                    if coords not in seen_lines and coords_rev not in seen_lines:
                                        seen_lines.add(coords)
                                        block_lines.append(LineString(coords))
                            elif be.dxftype() == 'LWPOLYLINE' and not be.closed:
                                pts = list(be.points())
                                if len(pts) >= 2:
                                    coords = tuple((round(p[0], 4), round(p[1], 4)) for p in pts)
                                    coords_rev = tuple(reversed(coords))
                                    if coords not in seen_lines and coords_rev not in seen_lines:
                                        seen_lines.add(coords)
                                        block_lines.append(LineString(coords))
                            elif be.dxftype() == 'LINE':
                                start = (round(be.dxf.start.x, 4), round(be.dxf.start.y, 4))
                                end = (round(be.dxf.end.x, 4), round(be.dxf.end.y, 4))
                                line_key = (start, end)
                                line_key_rev = (end, start)
                                if line_key not in seen_lines and line_key_rev not in seen_lines:
                                    seen_lines.add(line_key)
                                    block_lines.append(LineString([start, end]))
                            elif be.dxftype() == 'ARC':
                                # ARC를 선분들로 근사
                                try:
                                    center = (be.dxf.center.x, be.dxf.center.y)
                                    radius = be.dxf.radius
                                    start_angle = be.dxf.start_angle
                                    end_angle = be.dxf.end_angle
                                    import math
                                    # 각도 정규화
                                    if end_angle < start_angle:
                                        end_angle += 360
                                    angle_span = end_angle - start_angle
                                    num_segments = max(8, int(angle_span / 5))  # 최소 8개 세그먼트
                                    arc_pts = []
                                    for i in range(num_segments + 1):
                                        angle = math.radians(start_angle + (angle_span * i / num_segments))
                                        x = center[0] + radius * math.cos(angle)
                                        y = center[1] + radius * math.sin(angle)
                                        arc_pts.append((x, y))
                                    if len(arc_pts) >= 2:
                                        coords = tuple((round(x, 4), round(y, 4)) for x, y in arc_pts)
                                        coords_rev = tuple(reversed(coords))
                                        if coords not in seen_lines and coords_rev not in seen_lines:
                                            seen_lines.add(coords)
                                            block_lines.append(LineString(arc_pts))
                                except:
                                    pass

                        # 선분들을 합쳐서 폴리곤 생성
                        if block_lines:
                            try:
                                # 좌표 반올림으로 연결 오차 허용
                                rounded_lines = []
                                for line in block_lines:
                                    coords = list(line.coords)
                                    rounded_coords = [(round(x, 2), round(y, 2)) for x, y in coords]
                                    rounded_lines.append(LineString(rounded_coords))

                                polygons_from_lines = list(polygonize(rounded_lines))
                                for poly in polygons_from_lines:
                                    if not poly.is_valid:
                                        poly = poly.buffer(0)
                                    if poly.is_valid and poly.area > max_area:
                                        max_area = poly.area
                                        max_poly = poly
                            except Exception as e:
                                pass  # 폴리곤 생성 실패

                    # 원단명 기본값: 겉감
                    if not fabric_name:
                        fabric_name = "겉감"

                    # 패턴 이름이 없으면 PIECE NAME 번호 사용
                    if not pattern_name and piece_name:
                        pattern_name = piece_name

                    # pattern_group이 숫자만으로 구성된 경우 piece_name을 그룹으로 사용
                    # 예: 블록명 "1", "2", "3"... 인 경우 piece_name "BK1", "FRT2"를 그룹으로 사용
                    if piece_name and (not pattern_group or pattern_group.isdigit()):
                        pattern_group = piece_name

                    # 레이어 6의 LINE이 있으면 대칭선(Fold Line)으로 인식하여 미러링
                    if max_poly:
                        fold_line = None
                        for be in block:
                            if be.dxftype() == 'LINE':
                                layer = be.dxf.layer if hasattr(be.dxf, 'layer') else ''
                                if layer == '6':
                                    start = (be.dxf.start.x, be.dxf.start.y)
                                    end = (be.dxf.end.x, be.dxf.end.y)
                                    fold_line = (start, end)
                                    break

                        if fold_line:
                            # 대칭선을 기준으로 미러링
                            from shapely import affinity
                            start, end = fold_line

                            # 대칭선이 수평선인지 수직선인지 판단
                            dx = abs(end[0] - start[0])
                            dy = abs(end[1] - start[1])

                            if dx > dy:  # 수평선 (Y축 기준 미러링)
                                # 대칭선의 Y 좌표
                                mirror_y = (start[1] + end[1]) / 2
                                mirrored = affinity.scale(max_poly, xfact=1, yfact=-1, origin=(0, mirror_y))
                            else:  # 수직선 (X축 기준 미러링)
                                # 대칭선의 X 좌표
                                mirror_x = (start[0] + end[0]) / 2
                                mirrored = affinity.scale(max_poly, xfact=-1, yfact=1, origin=(mirror_x, 0))

                            # 원본과 미러링된 폴리곤 합치기 (약간 버퍼로 겹침 보장)
                            combined = max_poly.buffer(0.01).union(mirrored.buffer(0.01)).buffer(-0.01)
                            if combined.is_valid:
                                # MultiPolygon인 경우 가장 큰 폴리곤 선택
                                from shapely.geometry import MultiPolygon
                                if isinstance(combined, MultiPolygon):
                                    combined = max(combined.geoms, key=lambda g: g.area)
                                if isinstance(combined, Polygon):
                                    max_poly = combined
                                    max_area = combined.area

                    # 유효한 패턴만 추가 (너무 작은 패턴 제외)
                    if max_poly and max_area >= MIN_PATTERN_AREA:
                        # 선택된 사이즈 필터링 (selected_sizes가 지정된 경우)
                        if selected_sizes is not None:
                            # 사이즈명을 대문자로 비교
                            size_upper = size_name.upper() if size_name else ""
                            selected_upper = [s.upper() for s in selected_sizes]
                            if size_upper not in selected_upper:
                                continue  # 선택되지 않은 사이즈는 건너뛰기

                        # 내부선 추출 (외곽선 내부에 있는 LINE, POLYLINE, ARC, SPLINE 등)
                        interior_lines = []
                        # 외곽선 좌표 (비교용)
                        exterior_coords = set((round(x, 1), round(y, 1)) for x, y in max_poly.exterior.coords)

                        for be in block:
                            line_coords = None
                            try:
                                # LINE 엔티티
                                if be.dxftype() == 'LINE':
                                    start = (be.dxf.start.x, be.dxf.start.y)
                                    end = (be.dxf.end.x, be.dxf.end.y)
                                    line_coords = [start, end]

                                # POLYLINE (열린/닫힌 모두)
                                elif be.dxftype() == 'POLYLINE':
                                    pts = list(be.points())
                                    if len(pts) >= 2:
                                        line_coords = [(p[0], p[1]) for p in pts]
                                        if be.is_closed and len(line_coords) > 2:
                                            line_coords.append(line_coords[0])  # 닫힌 경우 시작점 추가

                                # LWPOLYLINE (열린/닫힌 모두)
                                elif be.dxftype() == 'LWPOLYLINE':
                                    pts = list(be.get_points(format='xy'))
                                    if len(pts) >= 2:
                                        line_coords = [(p[0], p[1]) for p in pts]
                                        if be.closed and len(line_coords) > 2:
                                            line_coords.append(line_coords[0])  # 닫힌 경우 시작점 추가

                                # ARC 엔티티 (호)
                                elif be.dxftype() == 'ARC':
                                    cx, cy = be.dxf.center.x, be.dxf.center.y
                                    r = be.dxf.radius
                                    start_angle = math.radians(be.dxf.start_angle)
                                    end_angle = math.radians(be.dxf.end_angle)
                                    # 호를 여러 점으로 분할
                                    if end_angle < start_angle:
                                        end_angle += 2 * math.pi
                                    num_points = max(10, int((end_angle - start_angle) / math.radians(10)))
                                    angles = [start_angle + (end_angle - start_angle) * i / num_points for i in range(num_points + 1)]
                                    line_coords = [(cx + r * math.cos(a), cy + r * math.sin(a)) for a in angles]

                                # SPLINE 엔티티
                                elif be.dxftype() == 'SPLINE':
                                    try:
                                        # 스플라인을 폴리라인으로 근사
                                        pts = list(be.flattening(0.5))  # 허용 오차 0.5
                                        if len(pts) >= 2:
                                            line_coords = [(p[0], p[1]) for p in pts]
                                    except:
                                        # flattening 실패 시 제어점 사용
                                        if hasattr(be, 'control_points'):
                                            pts = list(be.control_points)
                                            if len(pts) >= 2:
                                                line_coords = [(p[0], p[1]) for p in pts]

                                # CIRCLE 엔티티 (원)
                                elif be.dxftype() == 'CIRCLE':
                                    cx, cy = be.dxf.center.x, be.dxf.center.y
                                    r = be.dxf.radius
                                    num_points = 36
                                    angles = [2 * math.pi * i / num_points for i in range(num_points + 1)]
                                    line_coords = [(cx + r * math.cos(a), cy + r * math.sin(a)) for a in angles]

                                # ELLIPSE 엔티티 (타원)
                                elif be.dxftype() == 'ELLIPSE':
                                    try:
                                        pts = list(be.flattening(0.5))
                                        if len(pts) >= 2:
                                            line_coords = [(p[0], p[1]) for p in pts]
                                    except:
                                        pass
                            except:
                                continue

                            # 외곽선 내부에 있는 선만 추가 (외곽선 자체는 제외)
                            if line_coords and len(line_coords) >= 2:
                                # 외곽선과 동일한 좌표인지 확인
                                line_coords_rounded = set((round(x, 1), round(y, 1)) for x, y in line_coords)
                                overlap_ratio = len(line_coords_rounded & exterior_coords) / max(len(line_coords_rounded), 1)
                                if overlap_ratio > 0.8:  # 80% 이상 겹치면 외곽선 → 스킵
                                    continue

                                line_geom = LineString(line_coords)
                                # 선의 여러 지점이 외곽선 내부에 있는지 확인
                                check_points = [0.25, 0.5, 0.75]
                                inside_count = sum(1 for t in check_points if max_poly.contains(line_geom.interpolate(t, normalized=True)))
                                if inside_count >= 2:  # 3개 중 2개 이상이 내부에 있으면 추가
                                    interior_lines.append(line_coords)

                        # 그레인라인 감지 및 패턴 회전 (수직 정렬)
                        grainline_info = None  # (start, end) 좌표
                        rotation_applied = 0
                        grainline_angle, gl_start, gl_end = detect_grainline(block, return_coords=True, polygon=max_poly)
                        if grainline_angle is not None and gl_start and gl_end:
                            # 패턴 회전
                            centroid = (max_poly.centroid.x, max_poly.centroid.y)
                            max_poly, rotation_applied = rotate_polygon_to_vertical_grain(max_poly, grainline_angle)
                            # 그레인라인 좌표도 함께 회전
                            gl_start, gl_end = rotate_grainline_coords(gl_start, gl_end, rotation_applied, centroid)
                            grainline_info = (gl_start, gl_end)
                            # 내부선도 함께 회전
                            if interior_lines and abs(rotation_applied) > 0.1:
                                rotated_interior = []
                                for line_coords in interior_lines:
                                    rotated_line = []
                                    for px, py in line_coords:
                                        new_start, new_end = rotate_grainline_coords((px, py), (px, py), rotation_applied, centroid)
                                        rotated_line.append(new_start)
                                    rotated_interior.append(rotated_line)
                                interior_lines = rotated_interior

                        # 단위 스케일 적용 (인치 → cm 변환)
                        if unit_scale != 1.0:
                            from shapely import affinity
                            max_poly = affinity.scale(max_poly, xfact=unit_scale, yfact=unit_scale, origin=(0, 0))
                            if grainline_info:
                                gl_start, gl_end = grainline_info
                                gl_start = (gl_start[0] * unit_scale, gl_start[1] * unit_scale)
                                gl_end = (gl_end[0] * unit_scale, gl_end[1] * unit_scale)
                                grainline_info = (gl_start, gl_end)
                            # 내부선도 스케일 적용
                            if interior_lines:
                                scaled_interior = []
                                for line_coords in interior_lines:
                                    scaled_line = [(x * unit_scale, y * unit_scale) for x, y in line_coords]
                                    scaled_interior.append(scaled_line)
                                interior_lines = scaled_interior
                            max_area = max_poly.area

                        # 튜플: (poly, pattern_name, fabric_name, size_name, pattern_group, piece_name, dxf_quantity, grainline_info, interior_lines)
                        final.append((max_poly, pattern_name, fabric_name, size_name, pattern_group, piece_name, dxf_quantity, grainline_info, interior_lines))
                except Exception:
                    pass  # 블록 처리 실패

        # 방법 2: INSERT가 없으면 기존 방식 (레거시 DXF 호환)
        if not final:
            lines = []
            for e in msp:
                extract_lines(e, lines)

            rounded_lines = []
            for line in lines:
                coords = list(line.coords)
                rounded_coords = [(round(x, 1), round(y, 1)) for x, y in coords]
                rounded_lines.append(LineString(rounded_coords))

            merged_lines = linemerge(rounded_lines)
            raw_polys = list(polygonize(merged_lines))

            # 열린 선분 강제 닫기
            if hasattr(merged_lines, 'geoms'):
                chains = list(merged_lines.geoms)
            else:
                chains = [merged_lines] if merged_lines else []

            for chain in chains:
                if not chain.is_ring:
                    try:
                        start_pt = Point(chain.coords[0])
                        end_pt = Point(chain.coords[-1])
                        gap = start_pt.distance(end_pt)
                        if gap < 10.0:
                            closed_coords = list(chain.coords) + [chain.coords[0]]
                            new_poly = Polygon(closed_coords)
                            if new_poly.is_valid and new_poly.area > 0:
                                raw_polys.append(new_poly)
                    except:
                        pass

            candidates = [p for p in raw_polys if (p.area / 100) >= 10]
            candidates.sort(key=lambda x: x.area, reverse=True)

            # 레거시 방식에서만 중복 제거 (패턴 이름/원단명/사이즈/그룹 없음 → 기본값)
            added_polys = []
            for idx, p in enumerate(candidates):
                if not any(p.centroid.distance(e.centroid) < 50 for e in added_polys):
                    # 그레인라인 감지 및 패턴 회전 (수직 정렬) - 레거시 방식
                    grainline_angle = detect_grainline_for_polygon(msp, p)
                    grainline_info = None
                    if grainline_angle is not None:
                        centroid = (p.centroid.x, p.centroid.y)
                        p, _ = rotate_polygon_to_vertical_grain(p, grainline_angle)
                    # 단위 스케일 적용 (인치 → cm 변환)
                    if unit_scale != 1.0:
                        from shapely import affinity
                        p = affinity.scale(p, xfact=unit_scale, yfact=unit_scale, origin=(0, 0))

                    added_polys.append(p)
                    # 튜플: (poly, pattern_name, fabric_name, size_name, pattern_group, piece_name, dxf_qty, grainline_info, interior_lines)
                    final.append((p, "", "겉감", "", str(idx + 1), "", 0, grainline_info, []))

        # 면적 기준 정렬 (큰 것부터)
        final.sort(key=lambda x: x[0].area, reverse=True)
        return final, detected_base_size

    except Exception:
        return [], None


def sort_by_fabric():
    """
    세션 상태의 patterns와 df를 원단 우선으로 정렬합니다.
    패턴 수정(복사, 삭제, 원단 변경 등) 후 호출하여 일관된 정렬을 유지합니다.
    """
    if 'df' not in st.session_state or st.session_state.df.empty:
        return

    df = st.session_state.df
    patterns = st.session_state.patterns

    # 원단, 번호 순으로 정렬
    sort_indices = df.sort_values(by=['원단', '번호']).index.tolist()
    st.session_state.patterns = [patterns[i] for i in sort_indices]
    st.session_state.df = df.iloc[sort_indices].reset_index(drop=True)
    st.session_state.df["번호"] = range(1, len(st.session_state.df) + 1)

    # 체크박스 상태 초기화
    for i in range(len(st.session_state.patterns)):
        st.session_state[f"chk_{i}"] = False


def update_nesting_pattern_names():
    """
    네스팅 결과의 패턴 이름을 상세리스트의 구분명으로 업데이트합니다.
    상세리스트에서 구분 수정 시 호출하여 마카 표시에 반영합니다.
    pattern_id 형식: "df인덱스:이름\n사이즈_수량인덱스"
    """
    if 'nesting_results' not in st.session_state or not st.session_state.nesting_results:
        return
    if 'df' not in st.session_state or st.session_state.df.empty:
        return

    df = st.session_state.df

    # 네스팅 결과의 pattern_id 업데이트
    for fabric, result in st.session_state.nesting_results.items():
        if 'placements' not in result:
            continue
        for placement in result['placements']:
            old_id = placement['pattern_id']
            # pattern_id 형식: "df인덱스:이름\n사이즈_수량인덱스"
            # 예: "5:등판\nL_0" → df인덱스=5, 사이즈=L

            # 수량 인덱스 분리 (마지막 _숫자)
            if '_' in old_id:
                base_part, qty_idx = old_id.rsplit('_', 1)
            else:
                base_part, qty_idx = old_id, ""

            # 사이즈 분리
            if '\n' in base_part:
                idx_name_part, size_part = base_part.split('\n', 1)
            else:
                idx_name_part, size_part = base_part, ""

            # df 인덱스 추출
            if ':' in idx_name_part:
                df_idx_str = idx_name_part.split(':')[0]
            else:
                df_idx_str = idx_name_part

            # df에서 새 이름 가져오기
            try:
                df_idx = int(df_idx_str)
                if df_idx < len(df):
                    new_name = str(df.at[df_idx, '구분'])[:8] if df.at[df_idx, '구분'] else ""
                else:
                    new_name = ""
            except (ValueError, KeyError):
                new_name = ""

            # 새 pattern_id 생성
            if size_part:
                new_base = f"{df_idx_str}:{new_name}\n{size_part}"
            else:
                new_base = f"{df_idx_str}:{new_name}"

            if qty_idx:
                placement['pattern_id'] = f"{new_base}_{qty_idx}"
            else:
                placement['pattern_id'] = new_base


# ==============================================================================
# 4. UI 컴포넌트: 팝업 뷰어 (Dialog)
# ==============================================================================

@st.dialog("🔍 패턴 정밀 검토", width="large")
def show_detail_viewer(idx, pattern, fabric_name):
    """상세 보기 팝업창을 띄웁니다. (확대/이동/회전 기능 포함)"""
    st.caption("💡 사용법: **마우스 휠**로 줌, **드래그**로 이동, **슬라이더**로 회전")
    
    # 회전 컨트롤
    angle = st.slider("회전 각도 조절", 0, 360, 0, 90, label_visibility="collapsed")
    rotated_poly = affinity.rotate(pattern, angle, origin='centroid')
    
    # Plotly 데이터 준비
    x, y = rotated_poly.exterior.xy
    fill_color = get_fabric_color_hex(fabric_name)
    
    # 차트 그리기
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=list(x), y=list(y),
        fill="toself", fillcolor=fill_color,
        line=dict(color="black", width=2), mode='lines',
        name=f"Pattern {idx+1}"
    ))
    
    # 차트 레이아웃 설정 (CAD 스타일)
    fig.update_layout(
        xaxis=dict(visible=False), 
        yaxis=dict(visible=False, scaleanchor="x", scaleratio=1),
        plot_bgcolor='white', 
        margin=dict(l=10, r=10, t=10, b=10),
        height=600, 
        dragmode='pan' # 기본 도구를 '손바닥(이동)'으로 설정
    )
    st.plotly_chart(fig, width='stretch')
    
    # 하단 정보 표시
    minx, miny, maxx, maxy = rotated_poly.bounds
    w, h = (maxx - minx) / 10, (maxy - miny) / 10
    st.markdown(f"**📏 규격:** 가로 {w:.1f} cm x 세로 {h:.1f} cm")


# ==============================================================================
# 5. 메인 실행부 (Main Execution)
# ==============================================================================

# 세션 상태 초기화
if "df" not in st.session_state: st.session_state.df = None
if "patterns" not in st.session_state: st.session_state.patterns = None
if "loaded_file" not in st.session_state: st.session_state.loaded_file = None
if "dxf_sizes" not in st.session_state: st.session_state.dxf_sizes = None  # 스캔된 사이즈 목록
if "size_selection_done" not in st.session_state: st.session_state.size_selection_done = False  # 사이즈 선택 완료 여부
if "selected_load_sizes" not in st.session_state: st.session_state.selected_load_sizes = None  # 로딩할 사이즈 목록

# A. 파일 업로드 섹션
uploaded_file = st.file_uploader(
    "📁 DXF 파일을 여기에 드래그하거나 클릭하여 선택하세요",
    type=["dxf"],
    help="지원 형식: YUKA, Optitex, Gerber 등"
)

if uploaded_file is not None:
    # 파일이 변경되었으면 캐시 초기화 (파일명 + 크기 체크)
    file_key = f"{uploaded_file.name}_{uploaded_file.size}"
    if st.session_state.loaded_file != file_key:
        st.session_state.patterns = None
        st.session_state.df = None
        st.session_state.loaded_file = file_key
        st.session_state.dxf_sizes = None  # 사이즈 목록 초기화
        st.session_state.size_selection_done = False  # 사이즈 선택 상태 초기화
        st.session_state.selected_load_sizes = None  # 선택 사이즈 초기화
        # 체크박스 상태도 초기화
        for key in list(st.session_state.keys()):
            if key.startswith("chk_") or key.startswith("size_chk_"):
                del st.session_state[key]

    # 임시 파일 생성 (사이즈 스캔 및 패턴 로딩에 공용 사용)
    tmp_path = None
    if st.session_state.dxf_sizes is None or st.session_state.patterns is None:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".dxf") as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            tmp_path = tmp_file.name

    # 1단계: 사이즈 스캔 (파일 업로드 직후)
    if st.session_state.dxf_sizes is None and tmp_path:
        with st.spinner("사이즈 목록 스캔 중..."):
            scanned_sizes, scanned_base_size = scan_dxf_sizes(tmp_path)
            st.session_state.dxf_sizes = scanned_sizes
            st.session_state.dxf_base_size = scanned_base_size  # 기준사이즈 저장
            # 사이즈가 1개 이하면 바로 선택 완료 처리 (선택 UI 불필요)
            if len(scanned_sizes) <= 1:
                st.session_state.size_selection_done = True
                st.session_state.selected_load_sizes = scanned_sizes if scanned_sizes else None

    # 2단계: 사이즈 선택 UI (여러 사이즈가 있을 때)
    if st.session_state.dxf_sizes and len(st.session_state.dxf_sizes) > 1 and not st.session_state.size_selection_done:
        base_size = st.session_state.get('dxf_base_size')
        st.info(f"📏 **{len(st.session_state.dxf_sizes)}개 사이즈 발견** (기준: {base_size}) - 불러올 사이즈를 선택하세요")

        # 사이즈 체크박스 크기 확대 CSS
        st.markdown("""
        <style>
        .size-select-container label {
            font-size: 1.5rem !important;
            font-weight: bold !important;
        }
        .size-select-container .stCheckbox > label > div[data-testid="stMarkdownContainer"] > p {
            font-size: 1.5rem !important;
        }
        .size-select-container input[type="checkbox"] {
            transform: scale(1.5);
            margin-right: 10px;
        }
        </style>
        """, unsafe_allow_html=True)

        # 사이즈 체크박스 (가로 배열)
        st.markdown('<div class="size-select-container">', unsafe_allow_html=True)
        cols = st.columns(min(6, len(st.session_state.dxf_sizes)))
        for i, size in enumerate(st.session_state.dxf_sizes):
            col_idx = i % len(cols)
            with cols[col_idx]:
                # 기본값: 기준사이즈만 선택
                default_val = st.session_state.get(f"size_chk_{size}", size == base_size)
                st.checkbox(size, value=default_val, key=f"size_chk_{size}")
        st.markdown('</div>', unsafe_allow_html=True)

        # 선택 완료 버튼
        selected_count = sum(1 for size in st.session_state.dxf_sizes if st.session_state.get(f"size_chk_{size}", size == base_size))
        st.write(f"선택된 사이즈: **{selected_count}개** / 전체 {len(st.session_state.dxf_sizes)}개")

        if st.button(f"🚀 선택한 {selected_count}개 사이즈 불러오기", type="primary", use_container_width=True, disabled=(selected_count == 0)):
            # 선택된 사이즈 목록 저장
            selected = [size for size in st.session_state.dxf_sizes if st.session_state.get(f"size_chk_{size}", size == base_size)]
            st.session_state.selected_load_sizes = selected if selected else None
            st.session_state.size_selection_done = True
            st.rerun()

        # 임시 파일 삭제 (사이즈 선택 전에는 패턴 로딩 안 함)
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)
        st.stop()  # 사이즈 선택 완료 전까지 아래 코드 실행 안 함

    # 3단계: 선택된 사이즈만 패턴 로딩
    if st.session_state.patterns is None and st.session_state.size_selection_done:
        if tmp_path is None:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".dxf") as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                tmp_path = tmp_file.name

        with st.spinner("패턴 로딩 중..."):
            # 캐시 키용으로 리스트를 튜플로 변환
            sizes_tuple = tuple(st.session_state.selected_load_sizes) if st.session_state.selected_load_sizes else None
            patterns, detected_base_size = process_dxf(tmp_path, sizes_tuple)
            style_no = extract_style_no(tmp_path)

            # 기준사이즈가 선택되지 않은 경우, 기준사이즈 패턴만 별도로 로드하여 수량/원단 정보 수집
            base_size_quantities_cache = {}  # {pattern_group: quantity}
            base_size_fabrics_cache = {}  # {pattern_group: fabric_name}

            if detected_base_size and (sizes_tuple is None or detected_base_size not in sizes_tuple):
                # 기준사이즈가 선택되지 않았으므로 별도 로드
                base_patterns, _ = process_dxf(tmp_path, (detected_base_size,))
                for p_data in base_patterns:
                    pattern_group = p_data[4]
                    fabric_name = p_data[2]
                    dxf_quantity = p_data[6] if len(p_data) > 6 else 0
                    if pattern_group:
                        if dxf_quantity > 0:
                            base_size_quantities_cache[pattern_group] = dxf_quantity
                        if fabric_name:
                            base_size_fabrics_cache[pattern_group] = fabric_name

            st.session_state.base_size_quantities_cache = base_size_quantities_cache
            st.session_state.base_size_fabrics_cache = base_size_fabrics_cache

        os.remove(tmp_path)  # 임시 파일 삭제
        st.session_state.patterns = patterns
        st.session_state.detected_base_size = detected_base_size  # DXF에서 추출한 기준사이즈
        st.session_state.style_no = style_no
        st.session_state.original_pattern_count = len(patterns)  # 원본 패턴 수 저장
        st.session_state.thumbnail_cache = {}  # 썸네일 캐시 초기화
        
        # 패턴 DB 로드 (수량 추천용)
        pattern_db = None
        if PATTERN_DB_AVAILABLE:
            try:
                pattern_db = PatternDB()
            except:
                pass

        # 사이즈 목록 추출 (constants.size_sort_key 사용)
        all_sizes = sorted(set(p[3] for p in patterns if p[3]), key=size_sort_key)
        st.session_state.all_sizes = all_sizes

        # 기준사이즈 결정
        # 1순위: DXF에서 추출한 기준사이즈 (BASE_SIZE: 또는 REF_SIZE:)
        # 2순위: M 사이즈가 있으면 M 사용 (의류 산업 표준)
        # 3순위: 없으면 가운데 사이즈 사용
        if detected_base_size and detected_base_size in all_sizes:
            st.session_state.base_size = detected_base_size
        elif 'M' in all_sizes:
            st.session_state.base_size = 'M'
        elif all_sizes:
            mid_idx = len(all_sizes) // 2
            st.session_state.base_size = all_sizes[mid_idx]
        else:
            st.session_state.base_size = None

        # 선택 사이즈 초기화
        # 1순위: 로딩 시 선택한 사이즈 (다이얼로그에서 선택한 것)
        # 2순위: 기준사이즈만 선택
        selected_load_sizes = st.session_state.get('selected_load_sizes')
        if selected_load_sizes:
            # 로딩 시 선택한 사이즈를 그대로 사용 (all_sizes에 있는 것만)
            st.session_state.selected_sizes = [s for s in selected_load_sizes if s in all_sizes]
        elif st.session_state.base_size and st.session_state.base_size in all_sizes:
            st.session_state.selected_sizes = [st.session_state.base_size]
        elif all_sizes:
            st.session_state.selected_sizes = [all_sizes[0]]
        else:
            st.session_state.selected_sizes = []

        # 누락된 사이즈 패턴 자동 채우기
        # pattern_group별로 어떤 사이즈가 있는지 확인하고, 기준사이즈 패턴으로 채움
        if all_sizes and st.session_state.base_size:
            base_size = st.session_state.base_size

            # pattern_group별 사이즈 매핑
            group_size_map = {}  # {(pattern_group, fabric): {size: pattern_idx, ...}}
            for idx, p_data in enumerate(patterns):
                # 튜플: (poly, pattern_name, fabric_name, size_name, pattern_group, piece_name, dxf_quantity)
                pattern_group = p_data[4]
                size_name = p_data[3]
                fabric_name = p_data[2]
                if pattern_group and size_name:
                    key = (pattern_group, fabric_name if fabric_name else "겉감")
                    if key not in group_size_map:
                        group_size_map[key] = {}
                    group_size_map[key][size_name] = idx

            # 누락된 사이즈 찾아서 기준사이즈 패턴으로 채우기
            added_patterns = []
            missing_info = []
            for (pattern_group, fabric), size_dict in group_size_map.items():
                if base_size in size_dict:
                    # 기준사이즈 패턴이 있는 경우만 처리
                    base_idx = size_dict[base_size]
                    base_pattern = patterns[base_idx]

                    for size in all_sizes:
                        if size not in size_dict:
                            # 누락된 사이즈 - 기준사이즈 패턴 복사 (9개 요소)
                            new_pattern = (
                                base_pattern[0],  # poly (기준사이즈 형상 사용)
                                base_pattern[1],  # pattern_name
                                base_pattern[2],  # fabric_name
                                size,             # 새 사이즈
                                pattern_group,    # pattern_group
                                base_pattern[5] if len(base_pattern) > 5 else "",  # piece_name
                                base_pattern[6] if len(base_pattern) > 6 else 0,   # dxf_quantity
                                base_pattern[7] if len(base_pattern) > 7 else None, # grainline_info
                                base_pattern[8] if len(base_pattern) > 8 else []   # interior_lines
                            )
                            added_patterns.append(new_pattern)
                            missing_info.append(f"{pattern_group}_{size}")

            if added_patterns:
                patterns.extend(added_patterns)
                st.session_state.patterns = patterns
                st.toast(f"⚠️ 누락 사이즈 {len(added_patterns)}개 자동 추가: {', '.join(missing_info[:5])}{'...' if len(missing_info) > 5 else ''}")

        # 초기 데이터프레임 생성 (같은 패턴은 사이즈별 동일 번호)
        # 1단계: 패턴 정보 수집 및 번호 매핑 생성
        pattern_info = []
        pattern_number_map = {}  # (pattern_name, fabric) -> 번호
        current_number = 0

        # 기준사이즈 수량/원단 수집 (pattern_group별)
        # 기준사이즈에만 QUANTITY/CATEGORY가 있는 DXF 파일 지원
        # 핵심 규칙: 기준사이즈에 수량이 있으면 모든 사이즈가 따르고, 없으면 모든 사이즈가 추론 사용
        base_size = st.session_state.get('base_size') or st.session_state.get('detected_base_size')
        base_size_quantities = {}  # {pattern_group: quantity}
        base_size_fabrics = {}  # {pattern_group: fabric_name}
        base_size_has_any_quantity = False  # 기준사이즈에 수량이 하나라도 있는지
        base_size_has_any_fabric = False  # 기준사이즈에 원단이 하나라도 있는지

        # 1. 먼저 캐시된 기준사이즈 정보 로드 (기준사이즈가 선택되지 않은 경우를 위해)
        cached_quantities = st.session_state.get('base_size_quantities_cache', {})
        cached_fabrics = st.session_state.get('base_size_fabrics_cache', {})
        if cached_quantities:
            base_size_quantities.update(cached_quantities)
            base_size_has_any_quantity = True
        if cached_fabrics:
            base_size_fabrics.update(cached_fabrics)
            base_size_has_any_fabric = True

        # 2. 선택된 패턴에서 기준사이즈 정보 수집 (캐시 덮어쓰기)
        if base_size:
            for p_data in patterns:
                size_name = p_data[3]
                pattern_group = p_data[4]
                fabric_name = p_data[2]
                dxf_quantity = p_data[6] if len(p_data) > 6 else 0

                # 기준사이즈이고 pattern_group이 있는 경우 저장
                if size_name == base_size and pattern_group:
                    if dxf_quantity > 0:
                        base_size_quantities[pattern_group] = dxf_quantity
                        base_size_has_any_quantity = True
                    if fabric_name:
                        base_size_fabrics[pattern_group] = fabric_name
                        base_size_has_any_fabric = True

        for i, p_data in enumerate(patterns):
            # 튜플: (poly, pattern_name, fabric_name, size_name, pattern_group, piece_name, dxf_quantity)
            poly = p_data[0]
            pattern_name = p_data[1]
            fabric_name = p_data[2]
            size_name = p_data[3]
            pattern_group = p_data[4]
            piece_name = p_data[5] if len(p_data) > 5 else ""
            dxf_quantity = p_data[6] if len(p_data) > 6 else 0

            minx, miny, maxx, maxy = poly.bounds
            w, h = (maxx - minx) / 10, (maxy - miny) / 10

            # 원단 결정: 기준사이즈 기준으로 전체 일관성 유지
            # 핵심 규칙:
            # - 기준사이즈에 원단이 있으면 → 모든 사이즈가 기준사이즈 원단 사용
            # - 기준사이즈에 원단이 없으면 → 모든 사이즈가 기본값 "겉감" 사용

            if base_size_has_any_fabric:
                # 기준사이즈에 원단이 있는 경우: 기준사이즈 원단만 사용
                if pattern_group and pattern_group in base_size_fabrics:
                    extracted_fabric = base_size_fabrics[pattern_group]
                else:
                    # pattern_group이 기준사이즈에 없는 경우: 기본값
                    extracted_fabric = "겉감"
            else:
                # 기준사이즈에 원단이 없는 경우: 기본값 "겉감"
                extracted_fabric = "겉감"

            # 수량 결정: 기준사이즈 기준으로 전체 일관성 유지
            # 핵심 규칙:
            # - 기준사이즈에 수량이 있으면 → 모든 사이즈가 기준사이즈 수량 사용
            # - 기준사이즈에 수량이 없으면 → 모든 사이즈가 추론 사용 (DXF QUANTITY 무시)

            db_used = False

            if base_size_has_any_quantity:
                # 기준사이즈에 수량이 있는 경우: 기준사이즈 수량만 사용
                if pattern_group and pattern_group in base_size_quantities:
                    count = base_size_quantities[pattern_group]
                    default_desc = pattern_name if pattern_name else "확인"
                    db_used = True
                else:
                    # pattern_group이 기준사이즈에 없는 경우: 패턴 DB 예측 또는 형상 추론 사용
                    if pattern_db and len(pattern_db.records) > 0:
                        pred_qty, pred_cat, confidence, refs = pattern_db.predict_quantity(poly)
                        if confidence >= 0.5:
                            count = pred_qty
                            default_desc = pred_cat
                            db_used = True

            # 기준사이즈에 수량이 없거나, 위에서 수량을 결정하지 못한 경우: 추론 사용
            if not db_used:
                # 패턴 DB 예측 시도
                if pattern_db and len(pattern_db.records) > 0:
                    pred_qty, pred_cat, confidence, refs = pattern_db.predict_quantity(poly)
                    if confidence >= 0.5:
                        count = pred_qty
                        default_desc = pred_cat
                        db_used = True

                # 형상 기반 추론
                if not db_used:
                    is_symmetric, sym_reason = check_symmetry(poly)

                    # 1. 좌우대칭 + 가로≥50cm + 세로≥45cm → BACK, 1개
                    if sym_reason == "좌우대칭" and w >= 50 and h >= 45:
                        count = 1
                        default_desc = "BACK"
                    # 2. 좌우대칭 + 가로≥50cm + 세로≥20cm + 세로<45cm → BACK YOKE, 1개
                    elif sym_reason == "좌우대칭" and w >= 50 and h >= 20 and h < 45:
                        count = 1
                        default_desc = "BACK YOKE"
                    # 3. 가로≥25cm + 세로≥40cm + 세로직선(≥35cm) 1개 이상 → FRONT, 2개
                    elif w >= 25 and h >= 40 and check_vertical_straight_edge(poly, 35):
                        count = 2
                        default_desc = "FRONT"
                    # 3. 좌우대칭 + 가로≥45cm + 세로≤15cm + 가로직선 1개 → BACK YOKE HEM, 1개
                    elif sym_reason == "좌우대칭" and w >= 45 and h <= 15 and check_horizontal_straight_edge(poly):
                        count = 1
                        default_desc = "BACK YOKE HEM"
                    # 5. 좌우대칭 + 가로≥50cm + 세로≤10cm + 평행선(85%) → BACK BOTTOM, 1개
                    elif sym_reason == "좌우대칭" and w >= 50 and h <= 10 and check_parallel_edges(poly, 0.85):
                        count = 1
                        default_desc = "BACK BOTTOM"
                    # 6. 좌우대칭 + 가로≤25cm + 세로≤15cm → FLAP, 4개
                    elif sym_reason == "좌우대칭" and w <= 25 and h <= 15:
                        count = 4
                        default_desc = "FLAP"
                    # 7. 상하대칭 + 가로≤26cm + 세로≤12cm → SLEEVE TAB, 4개
                    elif sym_reason == "상하대칭" and w <= 26 and h <= 12:
                        count = 4
                        default_desc = "SLEEVE TAB"
                    # 8. 나머지 → 확인, 2개
                    else:
                        count = 2
                        default_desc = "확인"

            # 구분(패턴 이름) 결정: DXF 원본 부위명 우선
            if pattern_name:
                desc = pattern_name
            elif db_used and dxf_quantity == 0:
                desc = default_desc
            else:
                desc = default_desc

            # 패턴 키: pattern_group + fabric으로 동일 패턴 식별
            # pattern_group은 DXF 블록명에서 추출된 패턴 번호 (예: "1", "2", "3")
            if pattern_group:
                pattern_key = (pattern_group, extracted_fabric)
            else:
                # pattern_group이 없으면 pattern_name 사용
                pattern_key = (pattern_name or f"P{i}", extracted_fabric)
            if pattern_key not in pattern_number_map:
                current_number += 1
                pattern_number_map[pattern_key] = current_number

            pattern_info.append({
                'poly': poly, 'pattern_name': pattern_name, 'extracted_fabric': extracted_fabric,
                'size_name': size_name, 'w': w, 'h': h, 'count': count, 'desc': desc,
                'pattern_key': pattern_key
            })

        # 2단계: 데이터프레임 생성
        data_list = []
        for info in pattern_info:
            pattern_num = pattern_number_map[info['pattern_key']]
            data_list.append({
                "형상": poly_to_base64(info['poly'], get_fabric_color_hex(info['extracted_fabric'])),
                "번호": pattern_num, "사이즈": info['size_name'], "원단": info['extracted_fabric'],
                "구분": info['desc'], "수량": info['count'],
                "가로(cm)": round(info['w'], 1), "세로(cm)": round(info['h'], 1),
                "면적_raw": info['poly'].area / 1000000,
                "버퍼_상": 0, "버퍼_하": 0, "버퍼_좌": 0, "버퍼_우": 0  # 패턴별 상하좌우 버퍼 (mm)
            })
        st.session_state.df = pd.DataFrame(data_list)
        # 원단별 정렬 (기본 정렬) - patterns 리스트도 동기화
        if not st.session_state.df.empty:
            sort_indices = st.session_state.df.sort_values(by=['원단', '번호']).index.tolist()
            st.session_state.patterns = [st.session_state.patterns[i] for i in sort_indices]
            st.session_state.df = st.session_state.df.iloc[sort_indices].reset_index(drop=True)
            st.session_state.df["번호"] = range(1, len(st.session_state.df) + 1)  # 번호 순차 재설정
        # 체크박스 상태 초기화
        for i in range(len(patterns)): st.session_state[f"chk_{i}"] = False

    # 데이터 로드
    patterns = st.session_state.patterns
    df = st.session_state.df

    if patterns:
        # 썸네일 비율 고정용 Max값 계산 (현재 남아있는 패턴 기준)
        # 패턴 삭제 시 가장 큰 패턴 기준으로 자동 재설정됨
        max_dim = 0
        for p_data in patterns:
            p = p_data[0]
            minx, miny, maxx, maxy = p.bounds
            max_dim = max(max_dim, maxx - minx, maxy - miny)
        zoom_span = max_dim * 1.1 if max_dim > 0 else 100  # 기본값 설정

        # ----------------------------------------------------------------
        # A. 사이즈 선택 UI (그레이딩된 DXF용)
        # ----------------------------------------------------------------
        all_sizes = st.session_state.get('all_sizes', [])
        if len(all_sizes) >= 1:
            # 기준사이즈 정보 표시 (원본 DXF 기준사이즈 우선 사용)
            detected_base = st.session_state.get('detected_base_size')
            base_size = detected_base if detected_base else st.session_state.get('base_size')

            if len(all_sizes) == 1:
                # 사이즈가 1개일 때: 간단한 정보만 표시
                st.markdown(f"#### 📐 사이즈: **{all_sizes[0]}**")
                st.session_state.selected_sizes = all_sizes.copy()
            else:
                # 사이즈가 2개 이상일 때: 선택 UI 표시
                # 개별 사이즈 체크박스 (기본값: 기준사이즈만 선택)
                default_selected = [base_size] if base_size in all_sizes else all_sizes[:1]
                selected_sizes = st.session_state.get('selected_sizes', default_selected)

                # 선택된 사이즈 표시 (기준사이즈 포함)
                selected_display = ', '.join(selected_sizes) if selected_sizes else '없음'
                st.markdown(f"#### 📐 사이즈 선택 (기준: **{base_size}**, 선택: {selected_display})")

                # 사이즈 선택 체크박스
                size_cols = st.columns(min(len(all_sizes), 10))

                new_selected = []
                for idx, size in enumerate(all_sizes):
                    with size_cols[idx] if idx < len(size_cols) else size_cols[-1]:
                        if st.checkbox(size, value=(size in selected_sizes), key=f"sz_{size}"):
                            new_selected.append(size)

                if new_selected != selected_sizes:
                    st.session_state.selected_sizes = new_selected
                    st.rerun()

            # 중첩 시각화
            with st.expander("🔍 사이즈 중첩 비교", expanded=True):
                # 패턴 그룹별로 분류 (상세 리스트 번호 기준) - 복사 패턴 제외
                from collections import defaultdict
                original_count = st.session_state.get('original_pattern_count', len(patterns))
                pattern_groups = defaultdict(list)

                # 기준 사이즈 패턴의 순차 번호 매핑 (pattern_group -> list_idx)
                base_size = st.session_state.get('base_size')
                group_to_list_idx = {}
                list_idx = 0
                for idx, p_data in enumerate(patterns):
                    if idx >= original_count:
                        continue
                    size_name = p_data[3]
                    pattern_group = p_data[4]
                    # 기준 사이즈이거나 사이즈 없는 패턴만 번호 부여
                    if not all_sizes or size_name == base_size or not size_name:
                        list_idx += 1
                        if pattern_group:
                            group_to_list_idx[pattern_group] = list_idx

                for idx, p_data in enumerate(patterns):
                    # 복사된 패턴은 제외 (원본 패턴 수 이후 인덱스)
                    if idx >= original_count:
                        continue

                    # 튜플: (poly, pattern_name, fabric_name, size_name, pattern_group, ...)
                    size_name = p_data[3]
                    pattern_group = p_data[4]
                    pattern_name = p_data[1]
                    if size_name:  # 사이즈가 있는 패턴만
                        # 상세 리스트의 순차 번호 사용 (일괄 수정 도구와 동일)
                        if pattern_group and pattern_group in group_to_list_idx:
                            group_key = f"{group_to_list_idx[pattern_group]}"
                        elif pattern_group:
                            group_key = f"{pattern_group}"
                        elif pattern_name:
                            group_key = pattern_name
                        else:
                            group_key = "기타"
                        # 인덱스와 함께 저장 (수량 조회용)
                        pattern_groups[group_key].append((idx, p_data))

                if pattern_groups:
                    # 전역 최대 크기 계산 (모든 패턴에 동일 비율 적용)
                    global_max_dim = 0
                    for group_patterns in pattern_groups.values():
                        for idx, p_data in group_patterns:
                            poly = p_data[0]
                            minx, miny, maxx, maxy = poly.bounds
                            max_dim = max(maxx - minx, maxy - miny)
                            if max_dim > global_max_dim:
                                global_max_dim = max_dim

                    # 그룹 수에 따라 컬럼 조정 (최대 6열)
                    num_groups = len(pattern_groups)
                    num_cols = min(num_groups, 6)

                    # 그룹을 숫자 기준으로 정렬 (1, 2, 3... 순서로)
                    def sort_key(item):
                        name = item[0]
                        # 숫자로만 구성된 경우 숫자 정렬
                        if name.isdigit():
                            return (0, int(name))
                        # "패턴 N" 형식 (fallback)
                        if name.startswith("패턴 "):
                            try:
                                return (0, int(name.replace("패턴 ", "")))
                            except:
                                return (1, name)
                        return (2, name)

                    sorted_groups = sorted(pattern_groups.items(), key=sort_key)

                    # 행별로 그룹 표시
                    for row_start in range(0, len(sorted_groups), num_cols):
                        row_groups = sorted_groups[row_start:row_start + num_cols]
                        overlay_cols = st.columns(len(row_groups))

                        for col_idx, (group_name, group_patterns) in enumerate(row_groups):
                            with overlay_cols[col_idx]:
                                # 사이즈 목록 추출
                                sizes_in_group = [p_data[3] for idx, p_data in group_patterns if p_data[3]]
                                sizes_display = ','.join(sizes_in_group) if sizes_in_group else ''

                                # 기준사이즈 패턴의 수량 가져오기
                                quantity = None
                                for idx, p_data in group_patterns:
                                    if p_data[3] == base_size:  # 기준사이즈인 경우
                                        if idx < len(df):
                                            quantity = df.loc[idx, '수량'] if '수량' in df.columns else None
                                        break
                                # 기준사이즈가 없으면 첫 번째 패턴의 수량 사용
                                if quantity is None and group_patterns:
                                    first_idx = group_patterns[0][0]
                                    if first_idx < len(df) and '수량' in df.columns:
                                        quantity = df.loc[first_idx, '수량']

                                # 캡션: "3번 (S,M,L) - 수량: 2장"
                                if quantity is not None:
                                    st.caption(f"**{group_name}번** ({sizes_display}) - 수량: {quantity}장")
                                else:
                                    st.caption(f"**{group_name}번** ({sizes_display})")

                                # create_overlay_visualization에는 p_data만 전달
                                patterns_only = [p_data for idx, p_data in group_patterns]
                                fig = create_overlay_visualization(
                                    patterns_only,
                                    st.session_state.selected_sizes,
                                    all_sizes,
                                    global_max_dim,
                                    base_size
                                )
                                st.plotly_chart(fig, use_container_width=True, key=f"overlay_{row_start}_{col_idx}", config={
                                    'scrollZoom': True,  # 마우스 휠 확대/축소
                                    'displayModeBar': True,
                                    'modeBarButtonsToRemove': ['select2d', 'lasso2d', 'autoScale2d'],
                                    'displaylogo': False
                                })
                else:
                    st.info("사이즈 정보가 있는 패턴이 없습니다.")

            st.divider()

        # ----------------------------------------------------------------
        # B. 일괄 수정 도구 (Batch Edit Tools)
        # ----------------------------------------------------------------
        st.markdown("#### ✨ 일괄 수정 도구")
        tool_col1, tool_col2, tool_col3 = st.columns([2.5, 1, 1])

        # 패턴 그룹 매핑 (동일 패턴의 다른 사이즈 연결)
        all_sizes = st.session_state.get('all_sizes', [])
        selected_sizes = st.session_state.get('selected_sizes', all_sizes)
        # 기준사이즈: DXF에서 추출한 값 또는 가운데 사이즈
        base_size = st.session_state.get('base_size', selected_sizes[0] if selected_sizes else None)

        # group_to_indices: (pattern_group, 원단) 조합으로 구분
        # 복사된 패턴은 원단이 다르므로 원본과 별도 그룹으로 관리됨
        group_to_indices = {}  # {(pattern_group, fabric): {size: idx, ...}}
        base_indices_set = set()
        for idx, p_data in enumerate(patterns):
            size_name = p_data[3]
            pattern_group = p_data[4]
            fabric = st.session_state.df.loc[idx, '원단'] if idx < len(st.session_state.df) else ''
            if pattern_group:
                group_key = (pattern_group, fabric)
                if group_key not in group_to_indices:
                    group_to_indices[group_key] = {}
                group_to_indices[group_key][size_name] = idx
            # 기본 사이즈 인덱스
            if not all_sizes or size_name == base_size or not size_name:
                base_indices_set.add(idx)

        # 1. 전체 선택/해제/복사/삭제/회전/뒤집기
        with tool_col1:
            c1, c2, c3, c4, c5, c6, c7, c8 = st.columns(8)
            if c1.button("✅전체", width='stretch', help="기본 사이즈 전체 선택"):
                for i in base_indices_set: st.session_state[f"chk_{i}"] = True
                st.rerun()
            if c2.button("⬜해제", width='stretch', help="모든 선택 해제"):
                for i in range(len(patterns)): st.session_state[f"chk_{i}"] = False
                st.rerun()
            if c3.button("📋복사", width='stretch', help="선택 패턴 복사"):
                sel_indices = [i for i in base_indices_set if st.session_state.get(f"chk_{i}")]
                if sel_indices:
                    new_patterns = list(st.session_state.patterns)
                    new_df = st.session_state.df.copy()
                    selected_sizes = st.session_state.get('selected_sizes', [])
                    all_sizes = st.session_state.get('all_sizes', [])

                    # 선택한 패턴 + 모든 선택된 사이즈 확장하여 복사
                    expanded_indices = set()
                    for idx in sel_indices:
                        pattern_group = patterns[idx][4]
                        fabric = st.session_state.df.loc[idx, '원단'] if idx < len(st.session_state.df) else ''
                        group_key = (pattern_group, fabric)
                        if pattern_group and group_key in group_to_indices:
                            # 선택된 사이즈의 동일 패턴 모두 추가
                            for size_name, size_idx in group_to_indices[group_key].items():
                                if not selected_sizes or size_name in selected_sizes:
                                    expanded_indices.add(size_idx)
                        else:
                            expanded_indices.add(idx)

                    for idx in sorted(expanded_indices):
                        orig_fabric = new_df.iloc[idx]["원단"]
                        new_fabric = "복사_" + orig_fabric
                        new_color = get_fabric_color_hex(new_fabric)

                        orig_pattern = st.session_state.patterns[idx]
                        new_patterns.append(orig_pattern)
                        new_row = new_df.iloc[idx].copy()
                        new_row["번호"] = len(new_patterns)
                        new_row["원단"] = new_fabric
                        new_row["형상"] = poly_to_base64(orig_pattern[0], new_color)
                        new_df = pd.concat([new_df, pd.DataFrame([new_row])], ignore_index=True)

                    st.session_state.patterns = new_patterns
                    st.session_state.df = new_df
                    sort_by_fabric()  # 원단 우선 정렬
                    st.rerun()
            if c4.button("🗑삭제", width='stretch', help="선택 패턴 삭제"):
                sel_indices = [i for i in base_indices_set if st.session_state.get(f"chk_{i}")]
                if sel_indices:
                    selected_sizes = st.session_state.get('selected_sizes', [])
                    all_sizes = st.session_state.get('all_sizes', [])
                    new_df = st.session_state.df

                    # 선택한 패턴 + 모든 선택된 사이즈 확장하여 삭제
                    delete_indices = set()
                    for idx in sel_indices:
                        pattern_group = patterns[idx][4]
                        fabric = st.session_state.df.loc[idx, '원단'] if idx < len(st.session_state.df) else ''
                        group_key = (pattern_group, fabric)
                        if pattern_group and group_key in group_to_indices:
                            for size_name, size_idx in group_to_indices[group_key].items():
                                if not selected_sizes or size_name in selected_sizes:
                                    delete_indices.add(size_idx)
                        else:
                            delete_indices.add(idx)

                    keep_indices = [i for i in range(len(patterns)) if i not in delete_indices]
                    new_patterns = [st.session_state.patterns[i] for i in keep_indices]
                    new_df = st.session_state.df.iloc[keep_indices].copy()
                    new_df = new_df.reset_index(drop=True)
                    new_df["번호"] = range(1, len(new_df) + 1)
                    st.session_state.patterns = new_patterns
                    st.session_state.df = new_df
                    sort_by_fabric()  # 원단 우선 정렬
                    st.rerun()

            # 회전/뒤집기 버튼들 (c5~c8)
            def transform_selected_patterns(transform_func, update_dimensions=False):
                """선택된 패턴에 변환 함수 적용"""
                sel_indices = [i for i in base_indices_set if st.session_state.get(f"chk_{i}")]
                if not sel_indices:
                    return False
                selected_sizes = st.session_state.get('selected_sizes', [])

                # 선택한 패턴 + 모든 선택된 사이즈 확장
                expanded_indices = set()
                for idx in sel_indices:
                    pattern_group = patterns[idx][4]
                    fabric = st.session_state.df.loc[idx, '원단'] if idx < len(st.session_state.df) else ''
                    group_key = (pattern_group, fabric)
                    if pattern_group and group_key in group_to_indices:
                        for size_name, size_idx in group_to_indices[group_key].items():
                            if not selected_sizes or size_name in selected_sizes:
                                expanded_indices.add(size_idx)
                    else:
                        expanded_indices.add(idx)

                for idx in expanded_indices:
                    p_data = list(st.session_state.patterns[idx])
                    poly = p_data[0]
                    transformed_poly, transformed_interior = transform_func(poly, p_data[8] if len(p_data) > 8 else [])
                    p_data[0] = transformed_poly
                    if len(p_data) > 8:
                        p_data[8] = transformed_interior
                    st.session_state.patterns[idx] = tuple(p_data)

                    # 썸네일 갱신
                    current_fabric = st.session_state.df.at[idx, "원단"]
                    new_color = get_fabric_color_hex(current_fabric)
                    st.session_state.df.at[idx, "형상"] = poly_to_base64(transformed_poly, new_color)

                    # 가로/세로 업데이트 (회전 시)
                    if update_dimensions:
                        minx, miny, maxx, maxy = transformed_poly.bounds
                        st.session_state.df.at[idx, "가로(cm)"] = round((maxx - minx) / 10, 1)
                        st.session_state.df.at[idx, "세로(cm)"] = round((maxy - miny) / 10, 1)

                # 썸네일 캐시 클리어
                if 'thumbnail_cache' in st.session_state:
                    st.session_state.thumbnail_cache = {}
                return True

            def rotate_90(poly, interior_lines):
                """90도 회전"""
                rotated_poly = affinity.rotate(poly, 90, origin='centroid')
                if interior_lines:
                    cx, cy = poly.centroid.x, poly.centroid.y
                    rotated_interior = []
                    for line_coords in interior_lines:
                        rotated_line = [(cx - (y - cy), cy + (x - cx)) for x, y in line_coords]
                        rotated_interior.append(rotated_line)
                    return rotated_poly, rotated_interior
                return rotated_poly, []

            def rotate_180(poly, interior_lines):
                """180도 회전"""
                rotated_poly = affinity.rotate(poly, 180, origin='centroid')
                if interior_lines:
                    cx, cy = poly.centroid.x, poly.centroid.y
                    rotated_interior = []
                    for line_coords in interior_lines:
                        rotated_line = [(2 * cx - x, 2 * cy - y) for x, y in line_coords]
                        rotated_interior.append(rotated_line)
                    return rotated_poly, rotated_interior
                return rotated_poly, []

            def flip_y(poly, interior_lines):
                """Y축 뒤집기 (상하 반전)"""
                flipped_poly = affinity.scale(poly, xfact=1, yfact=-1, origin='centroid')
                if interior_lines:
                    cy = poly.centroid.y
                    flipped_interior = []
                    for line_coords in interior_lines:
                        flipped_line = [(x, 2 * cy - y) for x, y in line_coords]
                        flipped_interior.append(flipped_line)
                    return flipped_poly, flipped_interior
                return flipped_poly, []

            def flip_x(poly, interior_lines):
                """X축 뒤집기 (좌우 반전)"""
                flipped_poly = affinity.scale(poly, xfact=-1, yfact=1, origin='centroid')
                if interior_lines:
                    cx = poly.centroid.x
                    flipped_interior = []
                    for line_coords in interior_lines:
                        flipped_line = [(2 * cx - x, y) for x, y in line_coords]
                        flipped_interior.append(flipped_line)
                    return flipped_poly, flipped_interior
                return flipped_poly, []

            if c5.button("🔄90°", width='stretch', help="선택 패턴 90° 회전"):
                if transform_selected_patterns(rotate_90, update_dimensions=True):
                    st.rerun()

            if c6.button("🔁180°", width='stretch', help="선택 패턴 180° 회전"):
                if transform_selected_patterns(rotate_180, update_dimensions=False):
                    st.rerun()

            if c7.button("↕Y반전", width='stretch', help="선택 패턴 상하 반전"):
                if transform_selected_patterns(flip_y, update_dimensions=False):
                    st.rerun()

            if c8.button("↔X반전", width='stretch', help="선택 패턴 좌우 반전"):
                if transform_selected_patterns(flip_x, update_dimensions=False):
                    st.rerun()

        # 2. 원단명 변경 (선택 패턴의 모든 사이즈에 적용)
        with tool_col2:
            f1, f2 = st.columns([3, 1])
            new_fabric = f1.text_input("원단명", value="안감", label_visibility="collapsed")
            if f2.button("원단적용", width='stretch'):
                sel_indices = [i for i in base_indices_set if st.session_state.get(f"chk_{i}")]
                if sel_indices and new_fabric:
                    new_color = get_fabric_color_hex(new_fabric)
                    selected_sizes = st.session_state.get('selected_sizes', [])
                    all_sizes = st.session_state.get('all_sizes', [])

                    # 선택한 패턴 + 모든 선택된 사이즈 확장하여 원단 변경
                    expanded_indices = set()
                    for idx in sel_indices:
                        pattern_group = patterns[idx][4]
                        fabric = st.session_state.df.loc[idx, '원단'] if idx < len(st.session_state.df) else ''
                        group_key = (pattern_group, fabric)
                        if pattern_group and group_key in group_to_indices:
                            for size_name, size_idx in group_to_indices[group_key].items():
                                if not selected_sizes or size_name in selected_sizes:
                                    expanded_indices.add(size_idx)
                        else:
                            expanded_indices.add(idx)

                    for idx in expanded_indices:
                        st.session_state.df.at[idx, "원단"] = new_fabric
                        st.session_state.df.at[idx, "형상"] = poly_to_base64(patterns[idx][0], new_color)
                    sort_by_fabric()  # 원단 우선 정렬
                    st.rerun()

        # 3. 수량 변경 (선택 패턴의 모든 사이즈에 적용)
        with tool_col3:
            n1, n2 = st.columns([3, 1])
            new_count = n1.number_input("수량", min_value=0, value=2, label_visibility="collapsed")
            if n2.button("수량적용", width='stretch'):
                sel_indices = [i for i in base_indices_set if st.session_state.get(f"chk_{i}")]
                if sel_indices:
                    selected_sizes = st.session_state.get('selected_sizes', [])
                    all_sizes = st.session_state.get('all_sizes', [])

                    # 선택한 패턴 + 모든 선택된 사이즈 확장하여 수량 변경
                    expanded_indices = set()
                    for idx in sel_indices:
                        pattern_group = patterns[idx][4]
                        fabric = st.session_state.df.loc[idx, '원단'] if idx < len(st.session_state.df) else ''
                        group_key = (pattern_group, fabric)
                        if pattern_group and group_key in group_to_indices:
                            for size_name, size_idx in group_to_indices[group_key].items():
                                if not selected_sizes or size_name in selected_sizes:
                                    expanded_indices.add(size_idx)
                        else:
                            expanded_indices.add(idx)

                    for idx in expanded_indices:
                        st.session_state.df.at[idx, "수량"] = new_count
                    st.rerun()

        # 4. 버퍼 설정 (선택 패턴의 상하좌우 버퍼) - 숨김 처리
        # st.markdown("##### 📐 패턴별 버퍼 설정 (mm)")
        # buf_col1, buf_col2, buf_col3, buf_col4, buf_col5 = st.columns([1, 1, 1, 1, 1])
        # with buf_col1:
        #     buf_top = st.number_input("상", min_value=0, max_value=100, value=12, key="buf_top_input", help="상단 버퍼 (mm)")
        # with buf_col2:
        #     buf_bottom = st.number_input("하", min_value=0, max_value=100, value=12, key="buf_bottom_input", help="하단 버퍼 (mm)")
        # with buf_col3:
        #     buf_left = st.number_input("좌", min_value=0, max_value=100, value=12, key="buf_left_input", help="좌측 버퍼 (mm)")
        # with buf_col4:
        #     buf_right = st.number_input("우", min_value=0, max_value=100, value=12, key="buf_right_input", help="우측 버퍼 (mm)")
        # with buf_col5:
        #     if st.button("버퍼적용", width='stretch', help="선택 패턴에 상하좌우 버퍼 적용"):
        #         sel_indices = [i for i in base_indices_set if st.session_state.get(f"chk_{i}")]
        #         if sel_indices:
        #             selected_sizes = st.session_state.get('selected_sizes', [])
        #
        #             # 선택한 패턴 + 모든 선택된 사이즈 확장하여 버퍼 적용
        #             expanded_indices = set()
        #             for idx in sel_indices:
        #                 pattern_group = patterns[idx][4]
        #                 fabric = st.session_state.df.loc[idx, '원단'] if idx < len(st.session_state.df) else ''
        #                 group_key = (pattern_group, fabric)
        #                 if pattern_group and group_key in group_to_indices:
        #                     for size_name, size_idx in group_to_indices[group_key].items():
        #                         if not selected_sizes or size_name in selected_sizes:
        #                             expanded_indices.add(size_idx)
        #                 else:
        #                     expanded_indices.add(idx)
        #
        #             for idx in expanded_indices:
        #                 st.session_state.df.at[idx, "버퍼_상"] = buf_top
        #                 st.session_state.df.at[idx, "버퍼_하"] = buf_bottom
        #                 st.session_state.df.at[idx, "버퍼_좌"] = buf_left
        #                 st.session_state.df.at[idx, "버퍼_우"] = buf_right
        #             st.rerun()

        st.divider()

        # ----------------------------------------------------------------
        # C. 썸네일 그리드 (20 Columns Grid) - 기본 사이즈만 표시
        # ----------------------------------------------------------------
        all_sizes = st.session_state.get('all_sizes', [])
        selected_sizes = st.session_state.get('selected_sizes', all_sizes)
        # 기준사이즈: DXF에서 추출한 값 또는 가운데 사이즈
        base_size = st.session_state.get('base_size', selected_sizes[0] if selected_sizes else None)

        if base_size and all_sizes:
            # 기준사이즈 출처 표시
            detected = st.session_state.get('detected_base_size')
            source_info = "(원본)" if detected and detected == base_size else "(중간)"
            st.caption(f"💡 **{base_size}** {source_info} 사이즈 썸네일 (숫자 버튼으로 확대)")
        else:
            st.caption("💡 썸네일 아래 **[숫자 버튼]**을 누르면 확대 창이 열립니다.")

        # 기본 사이즈만 필터링 (썸네일용)
        filtered_patterns = []
        for orig_idx, p_data in enumerate(patterns):
            # 튜플: (poly, pattern_name, fabric_name, size_name, pattern_group, piece_name, dxf_qty, grainline_info)
            poly = p_data[0]
            pattern_name = p_data[1]
            fabric_name = p_data[2]
            size_name = p_data[3]
            pattern_group = p_data[4]
            grainline_info = p_data[7] if len(p_data) > 7 else None
            if not all_sizes:  # 사이즈 없는 DXF
                filtered_patterns.append((orig_idx, poly, pattern_name, fabric_name, size_name, pattern_group, grainline_info))
            elif size_name == base_size:  # 기본 사이즈만
                filtered_patterns.append((orig_idx, poly, pattern_name, fabric_name, size_name, pattern_group, grainline_info))
            elif not size_name:  # 사이즈 없는 패턴
                filtered_patterns.append((orig_idx, poly, pattern_name, fabric_name, size_name, pattern_group, grainline_info))

        cols_per_row = 20
        rows = math.ceil(len(filtered_patterns) / cols_per_row)

        for row in range(rows):
            cols = st.columns(cols_per_row)
            for col_idx in range(cols_per_row):
                list_idx = row * cols_per_row + col_idx
                if list_idx < len(filtered_patterns):
                    orig_idx, p, pattern_name, _, size_name, pattern_group, grainline_info = filtered_patterns[list_idx]
                    # 원단명은 df에서 가져오기 (일괄수정 반영)
                    current_fabric = st.session_state.df.at[orig_idx, "원단"] if orig_idx < len(st.session_state.df) else "겉감"
                    with cols[col_idx]:
                        # 캐싱된 썸네일 사용 (깜빡임 방지, 그레인라인 표시)
                        thumbnail_data = get_cached_thumbnail(orig_idx, p, current_fabric, zoom_span, grainline_info)
                        st.image(thumbnail_data, use_container_width=True)

                        # 팝업 호출 버튼 (순차 번호 - 상세 리스트와 동일)
                        btn_label = f"{list_idx + 1}"
                        if st.button(btn_label, key=f"btn_zoom_{orig_idx}", width='stretch'):
                            show_detail_viewer(orig_idx, p, current_fabric)

                        # 선택 체크박스
                        st.checkbox("선택", key=f"chk_{orig_idx}", label_visibility="collapsed")

        st.divider()

        # ----------------------------------------------------------------
        # D. 하단 작업창: 리스트 & 요척 결과 (Results)
        # ----------------------------------------------------------------
        col1, col2 = st.columns([3, 2])
        
        # [왼쪽] 상세 리스트 (Data Editor) - 기본 사이즈만 편집, 모든 사이즈에 적용
        with col1:
            st.markdown("#### 📝 상세 리스트")

            all_sizes = st.session_state.get('all_sizes', [])
            selected_sizes = st.session_state.get('selected_sizes', all_sizes)

            # 기준사이즈: DXF에서 추출한 값 또는 가운데 사이즈
            base_size = st.session_state.get('base_size', selected_sizes[0] if selected_sizes else None)

            # 패턴 그룹별 인덱스 매핑 (동일 패턴의 다른 사이즈 연결)
            # (pattern_group, 원단) 조합으로 구분 - 복사된 패턴은 다른 원단이므로 별도 그룹
            group_to_indices = {}  # {(pattern_group, fabric): {size: idx, ...}}
            for idx, p_data in enumerate(patterns):
                size_name = p_data[3]
                pattern_group = p_data[4]
                fabric = st.session_state.df.loc[idx, '원단'] if idx < len(st.session_state.df) else ''
                if pattern_group:
                    group_key = (pattern_group, fabric)
                    if group_key not in group_to_indices:
                        group_to_indices[group_key] = {}
                    group_to_indices[group_key][size_name] = idx

            # 기본 사이즈 인덱스만 추출 (편집용)
            base_indices = []
            for idx, p_data in enumerate(patterns):
                size_name = p_data[3]
                if not all_sizes:  # 사이즈 없는 DXF
                    base_indices.append(idx)
                elif size_name == base_size:
                    base_indices.append(idx)
                elif not size_name:  # 사이즈 없는 패턴
                    base_indices.append(idx)

            # 선택된 모든 사이즈의 인덱스 (요척 계산용) - DataFrame 기반
            all_filtered_indices = []
            for idx in range(len(st.session_state.df)):
                size_val = st.session_state.df.at[idx, '사이즈'] if '사이즈' in st.session_state.df.columns else ''
                if not all_sizes or not size_val or size_val in selected_sizes:
                    all_filtered_indices.append(idx)
            st.session_state.filtered_indices = all_filtered_indices

            # 기본 사이즈 표시 (출처 표시)
            if base_size and all_sizes:
                detected = st.session_state.get('detected_base_size')
                source_info = "(원본)" if detected and detected == base_size else "(중간)"
                st.caption(f"✏️ **{base_size}** {source_info} 사이즈 편집 → 모든 사이즈에 적용")

            # 데이터프레임 생성 (기본 사이즈만, 사이즈 열 숨김)
            display_df = st.session_state.df.iloc[base_indices].copy()

            # 버퍼 포함 면적 계산 (mm² -> m²)
            def calc_buffered_area(row):
                """버퍼 포함 면적 계산 (m² 단위)"""
                base_area = row['면적_raw']  # 원본 면적 (m²)
                w_mm = row['가로(cm)'] * 10  # 가로 (mm)
                h_mm = row['세로(cm)'] * 10  # 세로 (mm)
                buf_top = row.get('버퍼_상', 0) if '버퍼_상' in row else 0
                buf_bottom = row.get('버퍼_하', 0) if '버퍼_하' in row else 0
                buf_left = row.get('버퍼_좌', 0) if '버퍼_좌' in row else 0
                buf_right = row.get('버퍼_우', 0) if '버퍼_우' in row else 0

                # 버퍼 추가 면적 (mm²)
                # 상하: 가로 * 버퍼높이
                # 좌우: 세로 * 버퍼폭
                # 모서리: 버퍼_상*버퍼_좌 + 버퍼_상*버퍼_우 + 버퍼_하*버퍼_좌 + 버퍼_하*버퍼_우
                buffer_area_mm2 = (
                    w_mm * (buf_top + buf_bottom) +
                    h_mm * (buf_left + buf_right) +
                    buf_top * buf_left + buf_top * buf_right +
                    buf_bottom * buf_left + buf_bottom * buf_right
                )
                buffer_area_m2 = buffer_area_mm2 / 1_000_000
                return base_area + buffer_area_m2

            display_df["면적_버퍼포함"] = display_df.apply(calc_buffered_area, axis=1)
            display_df["면적(cm²)"] = (display_df["면적_버퍼포함"] * 10000).round(1)
            display_df = display_df.drop(columns=["면적_raw", "면적_버퍼포함"])
            # 사이즈 열 숨김 (사이즈선택 UI에서 이미 선택됨)
            if "사이즈" in display_df.columns:
                display_df = display_df.drop(columns=["사이즈"])
            # 버퍼 컬럼 숨김
            buffer_cols = ["버퍼_상", "버퍼_하", "버퍼_좌", "버퍼_우"]
            display_df = display_df.drop(columns=[c for c in buffer_cols if c in display_df.columns])
            display_df = display_df.reset_index(drop=True)
            display_df["번호"] = range(1, len(display_df) + 1)  # 번호 순차 재설정

            edited_df = st.data_editor(
                display_df,
                hide_index=True,
                width='stretch',
                num_rows="fixed",
                disabled=["형상", "번호", "가로(cm)", "세로(cm)", "면적(cm²)"],
                column_config={
                    "형상": st.column_config.ImageColumn(
                        "형상", help="패턴 미리보기", width="small"
                    )
                },
                height=735,
                key="editor_base"
            )

            # 변경사항 감지 및 모든 사이즈에 적용
            if edited_df is not None:
                selected_sizes = st.session_state.get('selected_sizes', [])
                all_sizes = st.session_state.get('all_sizes', [])
                any_change = False  # 변경 여부 플래그

                for i in range(len(edited_df)):
                    base_idx = base_indices[i] if i < len(base_indices) else i

                    # 변경 확인
                    old_fabric = st.session_state.df.at[base_idx, "원단"]
                    new_fabric = edited_df.at[i, "원단"]
                    old_qty = st.session_state.df.at[base_idx, "수량"]
                    new_qty = edited_df.at[i, "수량"]
                    old_cat = st.session_state.df.at[base_idx, "구분"]
                    new_cat = edited_df.at[i, "구분"]

                    # 버퍼 변경 확인
                    old_buf_top = st.session_state.df.at[base_idx, "버퍼_상"] if "버퍼_상" in st.session_state.df.columns else 0
                    new_buf_top = edited_df.at[i, "버퍼_상"] if "버퍼_상" in edited_df.columns else 0
                    old_buf_bottom = st.session_state.df.at[base_idx, "버퍼_하"] if "버퍼_하" in st.session_state.df.columns else 0
                    new_buf_bottom = edited_df.at[i, "버퍼_하"] if "버퍼_하" in edited_df.columns else 0
                    old_buf_left = st.session_state.df.at[base_idx, "버퍼_좌"] if "버퍼_좌" in st.session_state.df.columns else 0
                    new_buf_left = edited_df.at[i, "버퍼_좌"] if "버퍼_좌" in edited_df.columns else 0
                    old_buf_right = st.session_state.df.at[base_idx, "버퍼_우"] if "버퍼_우" in st.session_state.df.columns else 0
                    new_buf_right = edited_df.at[i, "버퍼_우"] if "버퍼_우" in edited_df.columns else 0

                    buf_change = (old_buf_top != new_buf_top or old_buf_bottom != new_buf_bottom or
                                  old_buf_left != new_buf_left or old_buf_right != new_buf_right)

                    has_change = (old_fabric != new_fabric or old_qty != new_qty or old_cat != new_cat or buf_change)

                    if has_change:
                        any_change = True
                        base_size = patterns[base_idx][3]
                        base_area = st.session_state.df.at[base_idx, "면적_raw"]

                        # 기본 사이즈 업데이트
                        st.session_state.df.at[base_idx, "원단"] = new_fabric
                        st.session_state.df.at[base_idx, "수량"] = new_qty
                        st.session_state.df.at[base_idx, "구분"] = new_cat
                        if old_fabric != new_fabric:
                            st.session_state.df.at[base_idx, "형상"] = poly_to_base64(patterns[base_idx][0], get_fabric_color_hex(new_fabric))

                        # 버퍼 업데이트
                        if "버퍼_상" in st.session_state.df.columns:
                            st.session_state.df.at[base_idx, "버퍼_상"] = new_buf_top
                            st.session_state.df.at[base_idx, "버퍼_하"] = new_buf_bottom
                            st.session_state.df.at[base_idx, "버퍼_좌"] = new_buf_left
                            st.session_state.df.at[base_idx, "버퍼_우"] = new_buf_right

                        # 동일 패턴의 다른 사이즈에도 적용 (pattern_group + 원단으로 매칭)
                        base_pattern_group = patterns[base_idx][4] if base_idx < len(patterns) else None
                        base_fabric = st.session_state.df.loc[base_idx, '원단'] if base_idx < len(st.session_state.df) else ''
                        group_key = (base_pattern_group, base_fabric)
                        if base_pattern_group and group_key in group_to_indices:
                            for size_name, j in group_to_indices[group_key].items():
                                if j == base_idx:
                                    continue
                                # 선택된 사이즈인 경우만 적용
                                if not all_sizes or not size_name or size_name in selected_sizes:
                                    st.session_state.df.at[j, "원단"] = new_fabric
                                    st.session_state.df.at[j, "수량"] = new_qty
                                    st.session_state.df.at[j, "구분"] = new_cat
                                    if old_fabric != new_fabric and j < len(patterns):
                                        st.session_state.df.at[j, "형상"] = poly_to_base64(patterns[j][0], get_fabric_color_hex(new_fabric))
                                    # 버퍼도 동일하게 적용
                                    if "버퍼_상" in st.session_state.df.columns:
                                        st.session_state.df.at[j, "버퍼_상"] = new_buf_top
                                        st.session_state.df.at[j, "버퍼_하"] = new_buf_bottom
                                        st.session_state.df.at[j, "버퍼_좌"] = new_buf_left
                                        st.session_state.df.at[j, "버퍼_우"] = new_buf_right

                        # 구분 변경 시 네스팅 결과의 패턴 이름도 업데이트
                        if old_cat != new_cat:
                            update_nesting_pattern_names()

                # 변경이 있었을 때만 한 번 리렌더링
                if any_change:
                    st.rerun()

        # [오른쪽] 요척 결과 카드 (Compact View) - 사이즈별 표시
        with col2:
            st.markdown("#### 📊 요척 결과")

            # 헤더 라벨 (가운데 정렬)
            h1, h2, h_u, h3, h4 = st.columns([1.4, 0.9, 0.9, 0.9, 1.4])
            h1.markdown("<div style='text-align: center; color: gray; font-size: 0.85rem;'>원단/사이즈</div>", unsafe_allow_html=True)
            h2.markdown("<div style='text-align: center; color: gray; font-size: 0.85rem;'>폭(W)</div>", unsafe_allow_html=True)
            h_u.markdown("<div style='text-align: center; color: gray; font-size: 0.85rem;'>단위</div>", unsafe_allow_html=True)
            h3.markdown("<div style='text-align: center; color: gray; font-size: 0.85rem;'>로스(%)</div>", unsafe_allow_html=True)
            h4.markdown("<div style='text-align: center; color: gray; font-size: 0.85rem;'>필요요척(YD)</div>", unsafe_allow_html=True)

            # 데이터 재계산 (필터링된 데이터 사용)
            filtered_indices = st.session_state.get('filtered_indices', list(range(len(st.session_state.df))))
            calc_df = st.session_state.df.iloc[filtered_indices].copy()

            # 버퍼 포함 면적 계산 함수
            def calc_buffered_area_for_yield(row):
                """요척 계산용 버퍼 포함 면적 (m² 단위)"""
                base_area = row['면적_raw']  # 원본 면적 (m²)
                w_mm = row['가로(cm)'] * 10  # 가로 (mm)
                h_mm = row['세로(cm)'] * 10  # 세로 (mm)
                buf_top = row.get('버퍼_상', 0) if '버퍼_상' in row else 0
                buf_bottom = row.get('버퍼_하', 0) if '버퍼_하' in row else 0
                buf_left = row.get('버퍼_좌', 0) if '버퍼_좌' in row else 0
                buf_right = row.get('버퍼_우', 0) if '버퍼_우' in row else 0

                # 버퍼 추가 면적 (mm²)
                buffer_area_mm2 = (
                    w_mm * (buf_top + buf_bottom) +
                    h_mm * (buf_left + buf_right) +
                    buf_top * buf_left + buf_top * buf_right +
                    buf_bottom * buf_left + buf_bottom * buf_right
                )
                buffer_area_m2 = buffer_area_mm2 / 1_000_000
                return base_area + buffer_area_m2

            # 버퍼 포함 면적 컬럼 추가
            calc_df['면적_버퍼포함'] = calc_df.apply(calc_buffered_area_for_yield, axis=1)

            # 선택된 사이즈 목록
            selected_sizes = st.session_state.get('selected_sizes', [])
            all_sizes_in_file = st.session_state.get('all_sizes', [])

            # 원단별 그룹
            fabric_groups = calc_df.groupby("원단")

            # 전체 합계 저장용
            total_yield_all = 0.0
            fabric_settings = {}  # 원단별 설정 저장

            fab_idx = 0
            for fabric_name, fabric_group in fabric_groups:
                color = get_fabric_color_hex(fabric_name)

                # 원단 헤더 (설정 입력)
                with st.container(border=True):
                    c1, c2, c_unit, c3, c4 = st.columns([1.4, 0.9, 0.9, 0.9, 1.4])

                    with c1:
                        st.markdown(f"""
                        <div style='background-color:{color}; padding:5px 0px; border-radius:4px; text-align:center;'>
                            <strong style='font-size:14px; color:#333;'>{fabric_name}</strong>
                        </div>""", unsafe_allow_html=True)

                    with c2:
                        input_width = st.number_input("W", value=58.00, min_value=10.0, step=0.1, format="%.2f", key=f"w_{fabric_name}", label_visibility="collapsed")

                    with c_unit:
                        unit = st.selectbox("U", ["in", "cm"], key=f"unit_{fabric_name}", label_visibility="collapsed")

                    with c3:
                        input_loss = st.number_input("L", value=15, min_value=0, key=f"l_{fabric_name}", label_visibility="collapsed")

                    # 설정 저장 (원단명 기반)
                    fabric_settings[fabric_name] = {
                        'width': input_width,
                        'unit': unit,
                        'loss': input_loss
                    }

                    # 원단 전체 합계 계산
                    if all_sizes_in_file and selected_sizes:
                        # 사이즈가 있는 경우: 선택된 사이즈만 합산
                        fabric_total_yd = 0.0
                        for size in selected_sizes:
                            size_data = fabric_group[fabric_group['사이즈'] == size]
                            if not size_data.empty:
                                size_area = sum(row['면적_버퍼포함'] * row['수량'] for _, row in size_data.iterrows())
                                if input_width > 0:
                                    width_m = input_width / 100 if unit == "cm" else (input_width * 2.54) / 100
                                    size_yd = ((size_area / width_m) / ((100-input_loss)/100)) * 1.09361
                                    fabric_total_yd += size_yd

                        with c4:
                            st.markdown(f"""
                            <div style='text-align:right; padding-top:5px;'>
                                <span style='font-size:16px; color:#0068c9; font-weight:bold;'>{fabric_total_yd:.2f} YD</span>
                            </div>""", unsafe_allow_html=True)

                        total_yield_all += fabric_total_yd
                    else:
                        # 사이즈 없는 경우: 전체 합산
                        group_area = sum(row['면적_버퍼포함'] * row['수량'] for _, row in fabric_group.iterrows())
                        if input_width > 0:
                            width_m = input_width / 100 if unit == "cm" else (input_width * 2.54) / 100
                            req_yd = ((group_area / width_m) / ((100-input_loss)/100)) * 1.09361
                        else:
                            req_yd = 0

                        with c4:
                            st.markdown(f"""
                            <div style='text-align:right; padding-top:5px;'>
                                <span style='font-size:18px; color:#0068c9; font-weight:bold;'>{req_yd:.2f} YD</span>
                            </div>""", unsafe_allow_html=True)

                        total_yield_all += req_yd

                # 사이즈별 상세 (사이즈가 있는 경우만)
                if all_sizes_in_file and selected_sizes:
                    for size in selected_sizes:
                        size_data = fabric_group[fabric_group['사이즈'] == size]
                        if not size_data.empty:
                            size_area = sum(row['면적_버퍼포함'] * row['수량'] for _, row in size_data.iterrows())

                            if input_width > 0:
                                width_m = input_width / 100 if unit == "cm" else (input_width * 2.54) / 100
                                size_yd = ((size_area / width_m) / ((100-input_loss)/100)) * 1.09361
                            else:
                                size_yd = 0

                            # 사이즈별 행 (들여쓰기)
                            s1, s2, s3, s4, s5 = st.columns([1.4, 0.9, 0.9, 0.9, 1.4])
                            with s1:
                                st.markdown(f"""
                                <div style='padding-left:20px; color:#666;'>
                                    └ {size}
                                </div>""", unsafe_allow_html=True)
                            with s5:
                                st.markdown(f"""
                                <div style='text-align:right; color:#666;'>
                                    {size_yd:.2f} YD
                                </div>""", unsafe_allow_html=True)

                fab_idx += 1

            # 전체 합계 표시
            st.markdown("---")
            t1, t2, t3, t4, t5 = st.columns([1.4, 0.9, 0.9, 0.9, 1.4])
            with t1:
                st.markdown("""
                <div style='padding:5px 0px; text-align:center;'>
                    <strong style='font-size:14px;'>합계</strong>
                </div>""", unsafe_allow_html=True)
            with t5:
                st.markdown(f"""
                <div style='text-align:right; padding-top:5px;'>
                    <span style='font-size:20px; color:#d94a4a; font-weight:bold;'>{total_yield_all:.2f} YD</span>
                </div>""", unsafe_allow_html=True)

            # 설정값 세션에 저장 (엑셀 내보내기용)
            st.session_state.fabric_settings = fabric_settings

            # ----------------------------------------------------------------
            # E. 엑셀 다운로드 버튼
            # ----------------------------------------------------------------
            st.divider()

            # 요척 결과 데이터 수집 (사이즈별)
            yield_data = []
            total_yield = 0.0

            # 설정값 가져오기
            fabric_settings = st.session_state.get('fabric_settings', {})
            selected_sizes = st.session_state.get('selected_sizes', [])
            all_sizes_in_file = st.session_state.get('all_sizes', [])

            for fabric_name, fabric_group in fabric_groups:
                settings = fabric_settings.get(fabric_name, {})
                input_width = settings.get('width', 58.0)
                unit = settings.get('unit', 'in')
                input_loss = settings.get('loss', 15)

                if all_sizes_in_file and selected_sizes:
                    # 사이즈별 행 추가
                    fabric_total = 0.0
                    for size in selected_sizes:
                        size_data = fabric_group[fabric_group['사이즈'] == size]
                        if not size_data.empty:
                            size_area = sum(row['면적_버퍼포함'] * row['수량'] for _, row in size_data.iterrows())
                            if input_width > 0:
                                width_m = input_width / 100 if unit == "cm" else (input_width * 2.54) / 100
                                size_yd = ((size_area / width_m) / ((100-input_loss)/100)) * 1.09361
                            else:
                                size_yd = 0

                            yield_data.append({
                                "원단명": fabric_name,
                                "사이즈": size,
                                "폭": input_width,
                                "단위": unit,
                                "효율(%)": 100 - input_loss,
                                "필요요척(YD)": round(size_yd, 2)
                            })
                            fabric_total += size_yd

                    # 원단 소계
                    yield_data.append({
                        "원단명": f"{fabric_name} 소계",
                        "사이즈": "",
                        "폭": "",
                        "단위": "",
                        "효율(%)": "",
                        "필요요척(YD)": round(fabric_total, 2)
                    })
                    total_yield += fabric_total
                else:
                    # 사이즈 없는 경우
                    group_area = sum(row['면적_버퍼포함'] * row['수량'] for _, row in fabric_group.iterrows())
                    if input_width > 0:
                        width_m = input_width / 100 if unit == "cm" else (input_width * 2.54) / 100
                        req_yd = ((group_area / width_m) / ((100-input_loss)/100)) * 1.09361
                    else:
                        req_yd = 0

                    yield_data.append({
                        "원단명": fabric_name,
                        "사이즈": "",
                        "폭": input_width,
                        "단위": unit,
                        "효율(%)": 100 - input_loss,
                        "필요요척(YD)": round(req_yd, 2)
                    })
                    total_yield += req_yd

            # 전체 합계 행 추가
            yield_data.append({
                "원단명": "합계",
                "사이즈": "",
                "폭": "",
                "단위": "",
                "효율(%)": "",
                "필요요척(YD)": round(total_yield, 2)
            })

            yield_df = pd.DataFrame(yield_data)

            # 엑셀 파일 생성
            excel_buffer = io.BytesIO()
            file_name = uploaded_file.name.replace('.dxf', '').replace('.DXF', '')
            style_no = st.session_state.get('style_no', '')

            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                # 시트1: 상세리스트 (선택된 모든 사이즈 포함)
                export_df = calc_df.copy()
                export_df["면적(cm²)"] = (export_df["면적_raw"] * 10000).round(1)
                detail_df = export_df.drop(columns=["형상", "면적_raw"], errors='ignore')
                # 파일명, 스타일번호 컬럼 추가
                detail_df.insert(0, "스타일번호", style_no)
                detail_df.insert(0, "파일명", file_name)
                detail_df.to_excel(writer, sheet_name='상세리스트', index=False)

                # 시트2: 요척결과 (사이즈별)
                # 파일명, 스타일번호 컬럼 추가
                yield_df.insert(0, "스타일번호", style_no)
                yield_df.insert(0, "파일명", file_name)
                yield_df.to_excel(writer, sheet_name='요척결과', index=False)

            excel_buffer.seek(0)

            st.download_button(
                label="📥 엑셀 다운로드",
                data=excel_buffer,
                file_name=f"{file_name}_요척결과.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch'
            )

        # ----------------------------------------------------------------
        # F. 네스팅 시뮬레이션 (원단별) - 전체 폭 사용
        # ----------------------------------------------------------------
        st.divider()
        st.markdown("#### 🧩 네스팅 시뮬레이션")

        # 원단별 설정
        fabric_list = st.session_state.df['원단'].unique().tolist()
        fabric_widths = {}
        marker_quantities = {}  # 원단별 벌수 (사이즈 없을 때 사용)
        size_quantities = {}    # 사이즈별 벌수
        target_efficiencies = {}

        # 사이즈 목록 확인
        all_sizes = st.session_state.get('all_sizes', [])
        selected_sizes = st.session_state.get('selected_sizes', all_sizes)
        has_multiple_sizes = len(selected_sizes) >= 2

        # 2컬럼 레이아웃: 왼쪽(공통설정) | 오른쪽(180도회전 + 원단별설정)
        left_col, right_col = st.columns(2)

        # 원단별 패턴 버퍼 저장용 딕셔너리
        fabric_buffers = {}

        with left_col:
            # 180도 회전 허용
            nest_rotation = st.checkbox(
                "180도 회전 허용",
                value=True,
                help="패턴을 180도 회전하여 배치",
                key="nest_rotation"
            )

            # 좌우 마주 보기 (좌우 미러링)
            nest_mirror = st.checkbox(
                "좌우 마주 보기",
                value=True,
                help="수량 2개 이상 패턴을 좌우 뒤집어서 마주보게 배치",
                key="nest_mirror"
            )

            # Sparrow 모드 (항상 활성화, UI 숨김)
            use_sparrow = SPARROW_AVAILABLE

            if use_sparrow:
                sparrow_time = st.number_input(
                    "최적화 시간(초)",
                    min_value=5, max_value=120, value=30,
                    help="더 긴 시간 = 더 좋은 결과",
                    key="sparrow_time"
                )
            else:
                sparrow_time = 30

            # 패턴 순서 옵션
            pattern_order = st.selectbox(
                "📐 패턴 순서",
                options=["default", "large_first", "small_first", "random"],
                format_func=lambda x: {"default": "기본순서", "large_first": "큰것부터", "small_first": "작은것부터", "random": "랜덤"}[x],
                index=0,
                help="패턴 입력 순서 (배치 결과에 영향)",
                key="pattern_order"
            )

            # 다중 시도 옵션
            multi_try = st.checkbox(
                "🎯 다중 시도 (최적 선택)",
                value=False,
                help="5가지 시드로 네스팅 후 최고 효율 선택",
                key="multi_try"
            )

            # 랜덤 시드 (다중 시도 비활성시만 표시)
            if not multi_try:
                nest_seed = st.number_input(
                    "🎲 랜덤 시드",
                    min_value=1, max_value=9999, value=42,
                    help="시드 변경 시 다른 배치 결과 생성",
                    key="nest_seed"
                )
            else:
                nest_seed = 42  # 다중 시도 시 내부에서 처리

            # 네스팅 실행 버튼
            run_nesting = st.button("🚀 네스팅 실행", width='stretch', type="primary")
            if 'nesting_elapsed' in st.session_state:
                st.caption(f"⏱️ {st.session_state.nesting_elapsed:.1f}초")

        with right_col:

            # 원단별 설정 헤더 (가운데 정렬)
            hcol1, hcol2, hcol3, hcol4, hcol5 = st.columns([2, 1, 1, 0.7, 1])
            with hcol1:
                st.markdown("<div style='text-align: center;'><b>원단</b></div>", unsafe_allow_html=True)
            with hcol2:
                st.markdown("<div style='text-align: center;'><b>효율%</b></div>", unsafe_allow_html=True)
            with hcol3:
                st.markdown("<div style='text-align: center;'><b>버퍼mm</b></div>", unsafe_allow_html=True)
            with hcol4:
                st.markdown("<div style='text-align: center;'><b>🔄90°</b></div>", unsafe_allow_html=True)
            with hcol5:
                st.markdown("<div style='text-align: center;'><b>벌수</b></div>", unsafe_allow_html=True)

            # 원단별 90도 회전 허용 딕셔너리
            fabric_90_rotations = {}

            # 원단별 설정 입력
            for i, fabric in enumerate(fabric_list):
                # 요척 결과에서 설정한 폭과 단위 가져오기 (원단명 기반 키)
                width_val = st.session_state.get(f"w_{fabric}", 58.0)
                unit_val = st.session_state.get(f"unit_{fabric}", "in")
                # cm로 변환
                if unit_val == "in":
                    width_cm = width_val * 2.54
                else:
                    width_cm = width_val
                fabric_widths[fabric] = width_cm

                col1, col2, col3, col4, col5 = st.columns([2, 1, 1, 0.7, 1])
                with col1:
                    st.markdown(f"<div style='text-align: center;'>{fabric}: {width_cm:.1f}cm</div>", unsafe_allow_html=True)
                with col2:
                    target_efficiencies[fabric] = st.number_input(
                        "효율%",
                        min_value=60, max_value=95, value=80,
                        key=f"target_eff_{i}",
                        label_visibility="collapsed"
                    )
                with col3:
                    # 원단별 패턴 버퍼
                    fabric_buffers[fabric] = st.number_input(
                        "버퍼",
                        min_value=0, max_value=50, value=0,
                        key=f"fabric_buffer_{i}",
                        label_visibility="collapsed",
                        help=f"{fabric} 원단의 패턴 버퍼 (둘레 확장)"
                    )
                with col4:
                    # 원단별 90도 회전 허용 (가운데 정렬)
                    _, chk_col, _ = st.columns([1, 2, 1])
                    with chk_col:
                        fabric_90_rotations[fabric] = st.checkbox(
                            "🔄",
                            value=False,
                            key=f"fabric_90rot_{i}",
                            label_visibility="collapsed",
                            help=f"{fabric} 원단 90° 회전 허용 (체크 시 0°/90°/180°/270° 회전 가능)"
                        )
                with col5:
                    marker_quantities[fabric] = st.number_input(
                        "벌수",
                        min_value=1, max_value=10, value=1,
                        key=f"marker_qty_{i}",
                        label_visibility="collapsed",
                        disabled=has_multiple_sizes  # 사이즈별 벌수 사용시 비활성화
                    )

            # 사이즈별 벌수 설정 (사이즈가 2개 이상일 때만 표시)
            if has_multiple_sizes:
                st.markdown("---")
                st.markdown("**📏 사이즈별 벌수**")

                # 사이즈별 벌수 입력 (가로 배치)
                size_cols = st.columns(len(selected_sizes))
                for si, size in enumerate(selected_sizes):
                    with size_cols[si]:
                        size_quantities[size] = st.number_input(
                            size,
                            min_value=0, max_value=10, value=1,
                            key=f"size_qty_{si}",
                            help=f"{size} 사이즈 벌수 (0=제외)"
                        )

        if run_nesting:
            import time
            start_time = time.time()
            spinner_msg = "🐦 Sparrow 최적화 중..." if use_sparrow else "원단별 네스팅 계산 중..."
            with st.spinner(spinner_msg):
                try:
                    nesting_results = {}

                    # 선택된 사이즈로 필터링된 인덱스 (상세 리스트와 동일)
                    all_sizes = st.session_state.get('all_sizes', [])
                    selected_sizes = st.session_state.get('selected_sizes', all_sizes)

                    filtered_indices_for_nesting = []
                    for idx, p_data in enumerate(patterns):
                        size_name = p_data[3]  # (poly, pattern_name, fabric_name, size_name, pattern_group)
                        if not all_sizes or not size_name or size_name in selected_sizes:
                            filtered_indices_for_nesting.append(idx)

                    # 원단별로 네스팅 실행
                    for fabric in fabric_list:
                        # 해당 원단 + 선택된 사이즈의 패턴만 필터링
                        fabric_indices = [
                            idx for idx in filtered_indices_for_nesting
                            if st.session_state.df.loc[idx, '원단'] == fabric
                        ]

                        if len(fabric_indices) == 0:
                            continue

                        # 패턴 데이터 수집
                        fabric_marker_qty = marker_quantities.get(fabric, 1)
                        pattern_data = []
                        total_qty_debug = 0  # 디버그: 총 수량 추적
                        for idx in fabric_indices:
                            if idx < len(patterns):
                                row = st.session_state.df.loc[idx]
                                poly = patterns[idx][0]
                                size_name = patterns[idx][3]  # 사이즈 정보
                                grainline_info = patterns[idx][7] if len(patterns[idx]) > 7 else None  # 그레인라인 정보
                                coords = list(poly.exterior.coords)[:-1]
                                coords_cm = [(p[0] / 10, p[1] / 10) for p in coords]

                                # 그레인라인 좌표도 cm로 변환
                                grainline_cm = None
                                if grainline_info:
                                    gl_start, gl_end = grainline_info
                                    grainline_cm = ((gl_start[0]/10, gl_start[1]/10), (gl_end[0]/10, gl_end[1]/10))

                                # 사이즈별 벌수 적용 (사이즈 2개 이상일 때)
                                if has_multiple_sizes and size_name:
                                    size_qty = size_quantities.get(size_name, 1)
                                    if size_qty == 0:  # 벌수 0이면 제외
                                        continue
                                    quantity = int(row['수량']) * size_qty
                                else:
                                    quantity = int(row['수량']) * fabric_marker_qty

                                total_qty_debug += quantity  # 디버그: 수량 누적

                                # 패턴ID: df인덱스(고유) + 이름(표시용) + 사이즈
                                # df인덱스는 정렬 후에도 유지되는 고유 식별자
                                pattern_name = str(row['구분'])[-10:] if row['구분'] else ""  # 표시용 이름 (뒤에서 10글자)
                                pattern_id = f"{idx}:{pattern_name}\n{size_name[:4]}" if size_name else f"{idx}:{pattern_name}"
                                # 패턴별 상하좌우 버퍼 (mm)
                                buf_top = row.get('버퍼_상', 0) if '버퍼_상' in row else 0
                                buf_bottom = row.get('버퍼_하', 0) if '버퍼_하' in row else 0
                                buf_left = row.get('버퍼_좌', 0) if '버퍼_좌' in row else 0
                                buf_right = row.get('버퍼_우', 0) if '버퍼_우' in row else 0

                                pattern_data.append({
                                    'coords_cm': coords_cm,
                                    'quantity': quantity,
                                    'pattern_id': pattern_id,
                                    'area_cm2': poly.area / 100,
                                    'df_idx': idx,  # 원본 df 인덱스 저장
                                    'grainline_cm': grainline_cm,  # 그레인라인 좌표 (cm)
                                    'buffer_top': buf_top,
                                    'buffer_bottom': buf_bottom,
                                    'buffer_left': buf_left,
                                    'buffer_right': buf_right
                                })

                        # 디버그: 상세 리스트 수량 합계 계산
                        df_fabric_indices = [i for i in range(len(st.session_state.df)) if st.session_state.df.loc[i, '원단'] == fabric]
                        df_qty_sum = sum(int(st.session_state.df.loc[i, '수량']) for i in df_fabric_indices)

                        # 사이즈별 벌수 적용한 예상 수량
                        if has_multiple_sizes:
                            expected_qty = df_qty_sum * sum(size_quantities.get(s, 1) for s in size_quantities if size_quantities.get(s, 1) > 0)
                        else:
                            expected_qty = df_qty_sum * fabric_marker_qty

                        st.info(f"📊 {fabric}: 상세리스트 {len(df_fabric_indices)}개(수량합:{df_qty_sum}), 네스팅 {len(pattern_data)}종(총:{total_qty_debug}개)")

                        if total_qty_debug != expected_qty and has_multiple_sizes:
                            st.warning(f"⚠️ 수량 불일치: 예상 {expected_qty}개, 실제 {total_qty_debug}개")

                        width_cm = fabric_widths[fabric]
                        # 원단별 패턴 버퍼
                        fabric_buffer = fabric_buffers.get(fabric, 0)

                        if use_sparrow and SPARROW_AVAILABLE:
                            # Sparrow 네스팅 (버퍼로 패턴 둘레 확장)
                            allow_90 = fabric_90_rotations.get(fabric, False)

                            if multi_try:
                                # 다중 시도: 5가지 시드로 실행 후 최고 효율 선택
                                best_result = None
                                best_efficiency = 0
                                best_seed = 42
                                test_seeds = [42, 123, 456, 789, 1024]
                                for test_seed in test_seeds:
                                    test_result = run_sparrow_nesting(
                                        pattern_data, width_cm, sparrow_time // 2, nest_rotation, 0, nest_mirror, fabric_buffer, allow_90, test_seed, pattern_order
                                    )
                                    if test_result.get('efficiency', 0) > best_efficiency:
                                        best_efficiency = test_result.get('efficiency', 0)
                                        best_result = test_result
                                        best_seed = test_seed
                                result = best_result
                                result['used_seed'] = best_seed
                            else:
                                # 단일 시도
                                result = run_sparrow_nesting(
                                    pattern_data, width_cm, sparrow_time, nest_rotation, 0, nest_mirror, fabric_buffer, allow_90, nest_seed, pattern_order
                                )
                                result['used_seed'] = nest_seed
                        else:
                            # 기본 네스팅 엔진 (버퍼 미지원, spacing으로 대체)
                            fabric_target_eff = target_efficiencies.get(fabric, 80)
                            engine = NestingEngine(
                                sheet_width=width_cm * 10,
                                spacing=fabric_buffer,  # 기본 엔진은 spacing으로 대체
                                target_efficiency=fabric_target_eff
                            )
                            for p in pattern_data:
                                engine.add_pattern(
                                    list(p['coords_cm']),
                                    quantity=p['quantity'],
                                    pattern_id=p['pattern_id']
                                )
                            rotations = [0, 180] if nest_rotation else [0]
                            result = engine.run(rotations=rotations)

                        result['fabric'] = fabric
                        result['width_cm'] = width_cm
                        result['marker_quantity'] = fabric_marker_qty
                        result['size_quantities'] = size_quantities if has_multiple_sizes else {}
                        result['has_multiple_sizes'] = has_multiple_sizes
                        result['buffer'] = fabric_buffer  # 원단별 패턴 버퍼 저장
                        result['allow_90'] = allow_90  # 원단별 90도 회전 허용 저장
                        result['pattern_order'] = pattern_order  # 패턴 순서 저장
                        nesting_results[fabric] = result

                        # 디버그: 배치 실패 패턴 확인
                        placed = result.get('placed_count', 0)
                        total = result.get('total_count', 0)
                        if placed < total:
                            st.error(f"❌ {fabric}: 배치 실패! 입력 {total}개 중 {placed}개만 배치됨 ({total - placed}개 누락)")

                    # 결과 저장 (작업일시 + 실행시간 추가)
                    from datetime import datetime
                    st.session_state.nesting_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
                    st.session_state.nesting_elapsed = time.time() - start_time
                    st.session_state.nesting_results = nesting_results
                    st.rerun()

                except Exception as e:
                    st.error(f"네스팅 오류: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())

        # 네스팅 결과 표시 (원단별 - 2열 배치)
        if 'nesting_results' in st.session_state and st.session_state.nesting_results:
            results = st.session_state.nesting_results
            fabric_list_results = list(results.keys())

            # 2열로 배치
            for i in range(0, len(fabric_list_results), 2):
                cols = st.columns(2)

                for j, col in enumerate(cols):
                    if i + j < len(fabric_list_results):
                        fabric = fabric_list_results[i + j]
                        result = results[fabric]
                        color = get_fabric_color_hex(fabric)
                        marker_qty = result.get('marker_quantity', 1)
                        timestamp = st.session_state.get('nesting_timestamp', '')

                        # 선택된 사이즈 정보 표시 (예: S/2,M/3=5벌)
                        selected_sizes = st.session_state.get('selected_sizes', [])
                        result_size_qty = result.get('size_quantities', {})
                        result_has_multi = result.get('has_multiple_sizes', False)

                        if selected_sizes and result_has_multi and result_size_qty:
                            # 사이즈별 벌수 표시
                            size_parts = ','.join([f"{s}/{result_size_qty.get(s, 1)}" for s in selected_sizes if result_size_qty.get(s, 1) > 0])
                            total_qty = sum(result_size_qty.get(s, 1) for s in selected_sizes if result_size_qty.get(s, 1) > 0)
                            size_info = f"{size_parts}={total_qty}벌"
                        elif selected_sizes:
                            size_parts = ','.join([f"{s}/{marker_qty}" for s in selected_sizes])
                            total_qty = len(selected_sizes) * marker_qty
                            size_info = f"{size_parts}={total_qty}벌"
                        else:
                            size_info = f"{marker_qty}벌"

                        with col:
                            with st.expander(f"📦 {fabric}({size_info}) - {timestamp}", expanded=True):
                                if result['success']:
                                    # 결과 메트릭 (5열: 패턴수, 원단폭, 마카길이, 요척, 효율)
                                    m1, m2, m3, m4, m5 = st.columns([1, 1, 1.2, 1, 0.8])
                                    m1.metric("패턴수", f"{result['placed_count']}/{result['total_count']}")
                                    m2.metric("원단폭", f"{result['width_cm']:.0f} cm")
                                    m3.metric("마카길이", f"{result['used_length_cm']:.1f} cm")
                                    # 요척 = 마카길이(YD) / 벌수 (사이즈별 벌수 합계 사용)
                                    if result_has_multi and result_size_qty:
                                        total_marker_qty = sum(result_size_qty.get(s, 1) for s in selected_sizes if result_size_qty.get(s, 1) > 0)
                                    else:
                                        total_marker_qty = marker_qty
                                    yield_per_set = result['used_length_yd'] / total_marker_qty if total_marker_qty > 0 else result['used_length_yd']
                                    m4.metric("요척", f"{yield_per_set:.2f} YD")
                                    m5.metric("효율", f"{result['efficiency']}%")

                                    # 시각화
                                    try:
                                        if result.get('sparrow_mode'):
                                            fig = create_sparrow_visualization(result, result['width_cm'])
                                        else:
                                            fig = create_nesting_visualization(result, result['width_cm'])
                                        if fig:
                                            st.pyplot(fig)
                                            plt.close(fig)
                                    except Exception as e:
                                        st.warning(f"시각화 오류: {str(e)}")

                                    # 재네스팅 옵션
                                    st.markdown("---")

                                    # 사이즈별 벌수 설정 (사이즈 2개 이상일 때)
                                    re_all_sizes = st.session_state.get('all_sizes', [])
                                    re_selected_sizes = st.session_state.get('selected_sizes', re_all_sizes)
                                    re_has_multi = len(re_selected_sizes) >= 2

                                    re_size_quantities = {}
                                    if re_has_multi:
                                        st.markdown("**사이즈별 벌수**")
                                        re_size_cols = st.columns(len(re_selected_sizes))
                                        for si, size in enumerate(re_selected_sizes):
                                            with re_size_cols[si]:
                                                re_size_quantities[size] = st.number_input(
                                                    size,
                                                    min_value=0, max_value=10, value=1,
                                                    key=f"re_sz_{fabric}_{size}_{i+j}",
                                                    help=f"{size} 벌수 (0=제외)"
                                                )

                                    # 벌수 변경 (사이즈 1개일 때만 표시)
                                    if not re_has_multi:
                                        re_col1, re_col2 = st.columns([1, 1])
                                        with re_col1:
                                            new_qty = st.number_input(
                                                "벌수 변경",
                                                min_value=1, max_value=10,
                                                value=marker_qty,
                                                key=f"re_qty_{fabric}_{i+j}"
                                            )
                                        with re_col2:
                                            re_nest_btn = st.button("🔄 재네스팅", key=f"re_nest_{fabric}_{i+j}", width='stretch')
                                    else:
                                        new_qty = 1  # 사이즈별 벌수 사용 시 기본값
                                        re_nest_btn = st.button("🔄 재네스팅", key=f"re_nest_{fabric}_{i+j}", width='stretch')

                                    if re_nest_btn:
                                            # 해당 원단만 재네스팅
                                            with st.spinner(f"🐦 {fabric} 재네스팅 중..."):
                                                try:
                                                    import time
                                                    start_time = time.time()

                                                    # 선택된 사이즈 + 해당 원단의 패턴만 필터링
                                                    all_sizes = st.session_state.get('all_sizes', [])
                                                    selected_sizes = st.session_state.get('selected_sizes', all_sizes)

                                                    fabric_indices = []
                                                    for idx, p_data in enumerate(patterns):
                                                        size_name = p_data[3]
                                                        if st.session_state.df.loc[idx, '원단'] == fabric:
                                                            if not all_sizes or not size_name or size_name in selected_sizes:
                                                                fabric_indices.append(idx)

                                                    # 패턴 데이터 수집
                                                    pattern_data = []
                                                    for idx in fabric_indices:
                                                        if idx < len(patterns):
                                                            row = st.session_state.df.loc[idx]
                                                            poly = patterns[idx][0]
                                                            size_name = patterns[idx][3]
                                                            grainline_info = patterns[idx][7] if len(patterns[idx]) > 7 else None
                                                            coords = list(poly.exterior.coords)[:-1]
                                                            coords_cm = [(p[0] / 10, p[1] / 10) for p in coords]

                                                            # 그레인라인 좌표도 cm로 변환
                                                            grainline_cm = None
                                                            if grainline_info:
                                                                gl_start, gl_end = grainline_info
                                                                grainline_cm = ((gl_start[0]/10, gl_start[1]/10), (gl_end[0]/10, gl_end[1]/10))

                                                            # 사이즈별 벌수 적용
                                                            if re_has_multi and size_name:
                                                                sz_qty = re_size_quantities.get(size_name, 1)
                                                                if sz_qty == 0:
                                                                    continue
                                                                quantity = int(row['수량']) * sz_qty
                                                            else:
                                                                quantity = int(row['수량']) * new_qty

                                                            base_id = str(row['구분'])[:12] if row['구분'] else f"P{idx+1}"
                                                            pattern_id = f"{base_id}\n{size_name[:4]}" if size_name else base_id

                                                            # 패턴별 상하좌우 버퍼 (mm)
                                                            buf_top = row.get('버퍼_상', 0) if '버퍼_상' in row else 0
                                                            buf_bottom = row.get('버퍼_하', 0) if '버퍼_하' in row else 0
                                                            buf_left = row.get('버퍼_좌', 0) if '버퍼_좌' in row else 0
                                                            buf_right = row.get('버퍼_우', 0) if '버퍼_우' in row else 0

                                                            pattern_data.append({
                                                                'coords_cm': coords_cm,
                                                                'quantity': quantity,
                                                                'pattern_id': pattern_id,
                                                                'area_cm2': poly.area / 100,
                                                                'grainline_cm': grainline_cm,
                                                                'buffer_top': buf_top,
                                                                'buffer_bottom': buf_bottom,
                                                                'buffer_left': buf_left,
                                                                'buffer_right': buf_right
                                                            })

                                                    width_cm = result['width_cm']
                                                    # 원단별 패턴 버퍼 (저장된 값 또는 기본값 0)
                                                    fabric_buffer = result.get('buffer', 0)

                                                    # Sparrow 네스팅 실행 (버퍼로 패턴 둘레 확장)
                                                    # 원단별 90도 회전 설정 가져오기 (저장된 값 사용)
                                                    allow_90 = result.get('allow_90', False)
                                                    used_seed = result.get('used_seed', st.session_state.get('nest_seed', 42))
                                                    used_order = st.session_state.get('pattern_order', 'default')
                                                    if SPARROW_AVAILABLE:
                                                        new_result = run_sparrow_nesting(
                                                            pattern_data, width_cm,
                                                            st.session_state.get('sparrow_time', 30),
                                                            st.session_state.get('nest_rotation', True),
                                                            0,
                                                            st.session_state.get('nest_mirror', False),
                                                            fabric_buffer,
                                                            allow_90,
                                                            used_seed,
                                                            used_order
                                                        )
                                                    else:
                                                        engine = NestingEngine(
                                                            sheet_width=width_cm * 10,
                                                            spacing=fabric_buffer,  # 기본 엔진은 spacing으로 대체
                                                            target_efficiency=80
                                                        )
                                                        for p in pattern_data:
                                                            engine.add_pattern(
                                                                list(p['coords_cm']),
                                                                quantity=p['quantity'],
                                                                pattern_id=p['pattern_id']
                                                            )
                                                        rotations = [0, 180] if st.session_state.get('nest_rotation', True) else [0]
                                                        new_result = engine.run(rotations=rotations)

                                                    new_result['fabric'] = fabric
                                                    new_result['width_cm'] = width_cm
                                                    new_result['marker_quantity'] = new_qty
                                                    new_result['size_quantities'] = re_size_quantities if re_has_multi else {}
                                                    new_result['has_multiple_sizes'] = re_has_multi
                                                    new_result['buffer'] = fabric_buffer  # 원단별 패턴 버퍼 유지

                                                    # 결과 업데이트
                                                    st.session_state.nesting_results[fabric] = new_result
                                                    st.session_state.nesting_elapsed = time.time() - start_time
                                                    from datetime import datetime
                                                    st.session_state.nesting_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
                                                    st.rerun()

                                                except Exception as e:
                                                    st.error(f"재네스팅 오류: {str(e)}")
                                else:
                                    st.warning(f"{fabric}: 패턴을 배치할 수 없습니다.")

            # 자동 최적화 버튼 (효율 70% 미만 원단 자동 벌수 조정)
            st.markdown("---")
            low_eff_fabrics = [f for f, r in results.items() if r.get('success') and r.get('efficiency', 0) < 70]

            opt_col1, opt_col2 = st.columns([1, 1])
            with opt_col1:
                if low_eff_fabrics:
                    st.caption(f"⚠️ 효율 70% 미만: {', '.join(low_eff_fabrics)}")
                else:
                    st.caption("✅ 모든 원단 효율 70% 이상")

            with opt_col2:
                if st.button("🎯 자동 최적화", width='stretch', disabled=len(low_eff_fabrics)==0):
                    with st.spinner("🔄 효율 70% 미만 원단 자동 최적화 중..."):
                        try:
                            import time
                            start_time = time.time()
                            optimized_count = 0

                            for fabric in low_eff_fabrics:
                                original_result = results[fabric]
                                width_cm = original_result['width_cm']
                                original_qty = original_result.get('marker_quantity', 1)
                                best_result = original_result
                                best_efficiency = original_result.get('efficiency', 0)
                                best_qty = original_qty

                                # 선택된 사이즈 + 해당 원단의 패턴 데이터 준비
                                all_sizes = st.session_state.get('all_sizes', [])
                                selected_sizes = st.session_state.get('selected_sizes', all_sizes)

                                fabric_indices = []
                                for idx, p_data in enumerate(patterns):
                                    size_name = p_data[3]
                                    if st.session_state.df.loc[idx, '원단'] == fabric:
                                        if not all_sizes or not size_name or size_name in selected_sizes:
                                            fabric_indices.append(idx)

                                base_pattern_data = []
                                for idx in fabric_indices:
                                    if idx < len(patterns):
                                        row = st.session_state.df.loc[idx]
                                        poly = patterns[idx][0]
                                        size_name = patterns[idx][3]
                                        grainline_info = patterns[idx][7] if len(patterns[idx]) > 7 else None
                                        coords = list(poly.exterior.coords)[:-1]
                                        coords_cm = [(p[0] / 10, p[1] / 10) for p in coords]

                                        # 그레인라인 좌표도 cm로 변환
                                        grainline_cm = None
                                        if grainline_info:
                                            gl_start, gl_end = grainline_info
                                            grainline_cm = ((gl_start[0]/10, gl_start[1]/10), (gl_end[0]/10, gl_end[1]/10))

                                        base_id = str(row['구분'])[:12] if row['구분'] else f"P{idx+1}"
                                        pattern_id = f"{base_id}\n{size_name[:4]}" if size_name else base_id

                                        # 패턴별 상하좌우 버퍼 (mm)
                                        buf_top = row.get('버퍼_상', 0) if '버퍼_상' in row else 0
                                        buf_bottom = row.get('버퍼_하', 0) if '버퍼_하' in row else 0
                                        buf_left = row.get('버퍼_좌', 0) if '버퍼_좌' in row else 0
                                        buf_right = row.get('버퍼_우', 0) if '버퍼_우' in row else 0

                                        base_pattern_data.append({
                                            'coords_cm': coords_cm,
                                            'base_quantity': int(row['수량']),
                                            'pattern_id': pattern_id,
                                            'area_cm2': poly.area / 100,
                                            'grainline_cm': grainline_cm,
                                            'buffer_top': buf_top,
                                            'buffer_bottom': buf_bottom,
                                            'buffer_left': buf_left,
                                            'buffer_right': buf_right
                                        })

                                # 원단별 패턴 버퍼 (저장된 값 또는 기본값 0)
                                fabric_buffer = result.get('buffer', 0)

                                # 벌수 2~5까지 시도하여 최적 효율 찾기
                                for try_qty in range(2, 6):
                                    pattern_data = []
                                    for p in base_pattern_data:
                                        pattern_data.append({
                                            'coords_cm': p['coords_cm'],
                                            'quantity': p['base_quantity'] * try_qty,
                                            'pattern_id': p['pattern_id'],
                                            'area_cm2': p['area_cm2'],
                                            'grainline_cm': p.get('grainline_cm'),
                                            'buffer_top': p.get('buffer_top', 0),
                                            'buffer_bottom': p.get('buffer_bottom', 0),
                                            'buffer_left': p.get('buffer_left', 0),
                                            'buffer_right': p.get('buffer_right', 0)
                                        })

                                    # 네스팅 실행 (버퍼로 패턴 둘레 확장)
                                    # 원단별 90도 회전 설정 가져오기 (저장된 값 사용)
                                    allow_90 = original_result.get('allow_90', False)
                                    used_seed = original_result.get('used_seed', st.session_state.get('nest_seed', 42))
                                    used_order = st.session_state.get('pattern_order', 'default')
                                    if SPARROW_AVAILABLE:
                                        test_result = run_sparrow_nesting(
                                            pattern_data, width_cm,
                                            st.session_state.get('sparrow_time', 30),
                                            st.session_state.get('nest_rotation', True),
                                            0,
                                            st.session_state.get('nest_mirror', False),
                                            fabric_buffer,
                                            allow_90,
                                            used_seed,
                                            used_order
                                        )
                                    else:
                                        engine = NestingEngine(
                                            sheet_width=width_cm * 10,
                                            spacing=fabric_buffer,  # 기본 엔진은 spacing으로 대체
                                            target_efficiency=80
                                        )
                                        for p in pattern_data:
                                            engine.add_pattern(
                                                list(p['coords_cm']),
                                                quantity=p['quantity'],
                                                pattern_id=p['pattern_id']
                                            )
                                        rotations = [0, 180] if st.session_state.get('nest_rotation', True) else [0]
                                        test_result = engine.run(rotations=rotations)

                                    test_eff = test_result.get('efficiency', 0)

                                    # 더 좋은 효율이면 저장
                                    if test_eff > best_efficiency:
                                        best_efficiency = test_eff
                                        best_result = test_result
                                        best_qty = try_qty

                                    # 목표 효율(80%) 이상이면 중단
                                    if test_eff >= 80:
                                        break

                                # 최적 결과로 업데이트
                                if best_qty != original_qty:
                                    best_result['fabric'] = fabric
                                    best_result['width_cm'] = width_cm
                                    best_result['marker_quantity'] = best_qty
                                    st.session_state.nesting_results[fabric] = best_result
                                    optimized_count += 1

                            st.session_state.nesting_elapsed = time.time() - start_time
                            from datetime import datetime
                            st.session_state.nesting_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

                            if optimized_count > 0:
                                st.success(f"✅ {optimized_count}개 원단 최적화 완료!")
                                st.rerun()
                            else:
                                st.info("최적화할 원단이 없습니다.")

                        except Exception as e:
                            st.error(f"자동 최적화 오류: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())

            # 네스팅 결과 엑셀 내보내기
            st.markdown("---")
            # 제목 정보 수집
            excel_style_no = st.session_state.get('style_no') or uploaded_file.name.replace('.dxf', '').replace('.DXF', '')
            excel_selected_sizes = st.session_state.get('selected_sizes', [])
            excel_base_size = st.session_state.get('detected_base_size') or st.session_state.get('base_size')
            excel_data = export_nesting_to_excel(
                results,
                st.session_state.get('nesting_timestamp', ''),
                style_no=excel_style_no,
                selected_sizes=excel_selected_sizes,
                base_size=excel_base_size
            )
            if excel_data:
                # DXF 파일 이름에서 확장자 제거
                dxf_base_name = uploaded_file.name.replace('.dxf', '').replace('.DXF', '')
                st.download_button(
                    label="📥 네스팅 결과 엑셀 다운로드",
                    data=excel_data,
                    file_name=f"{dxf_base_name}_네스팅결과.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width='stretch'
                )

    else:
        st.info("💡 DXF 파일을 업로드하면 패턴 분석이 시작됩니다.")

else:
    # 초기 화면 (파일 업로드 전)
    # 안내 문구
    st.markdown('''
    <div style="text-align: center; padding: 20px; background: #fffbe6; border-radius: 10px; border: 1px solid #ffe58f;">
        <p style="font-size: 16px; color: #d48806; font-weight: bold; margin: 0;">
            ⬆️ 위의 "Browse files" 버튼을 클릭하거나, 그 영역에 DXF 파일을 드래그하세요
        </p>
    </div>
    ''', unsafe_allow_html=True)

    # 50% 크기로 가운데 배치
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("#### 📺 사용 가이드")
        # YouTube 썸네일 + 링크 버튼
        st.image("https://img.youtube.com/vi/_-iXXdf-gMA/maxresdefault.jpg", width='stretch')
        st.link_button("▶️ YouTube에서 영상 보기", "https://youtu.be/_-iXXdf-gMA", width='stretch')